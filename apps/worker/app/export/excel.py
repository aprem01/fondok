# ruff: noqa: RUF001
"""Excel acquisition model builder.

Builds an institutional underwriting workbook from an engine outputs
dict (see ``evals/golden-set/kimpton-angler/expected/model.json``).

──────────────────────────────────────────────────────────────────────
Wave 4 W4.2 — Excel acquisition model refresh.

The legacy export (pre-Wave-2) shipped a fixed 10-sheet workbook with
none of the Wave 2/3 analytical artifacts (segment mix, PIP
displacement, three-bucket capex, op-ratio precedence, pricing
sensitivity grid, max-price solver, comparable sales, historical
baseline, STR forecast, named scenarios, LOI draft). Analysts who
exported the .xlsx saw a stale model.

This module now pulls every Wave 2/3 artifact into the workbook. Each
new sheet is **conditional** — it renders only when its data is
present on the ``model`` dict, so a barebones deal still produces a
clean model. The cover sheet's "Sections Included" list reflects what
actually shipped.

Sheet catalog (in build order — only those with data are included):

  Always (legacy):
    1.  Cover                       (refreshed)
    2.  Assumptions
    3.  Sources & Uses
    4.  Operating Proforma
    5.  Debt Schedule
    6.  Returns
    7.  Partnership
    8.  Variance
    9.  Market Comps

  Wave 2/3 (conditional):
    10. Revenue Mix                 — when segments_by_year set
    11. Renovation Plan             — when pip_displacement set
    12. Capital Plan (3-Bucket)     — when capex_schedule set
    13. Op-Ratio Provenance         — when op_ratio_provenance set
    14. Pricing Sensitivity         — when sensitivity_grid set
    15. Comparable Sales            — when comp_sales set
    16. Historical Baseline         — when historical_baseline set
    17. STR Forecast                — when str_forecast set
    18. Named Scenarios             — when named_scenarios set
    19. LOI Appendix                — when loi_draft set

Backward compat: every new sheet builder no-ops when its source data is
None / empty. The legacy 9-sheet workbook is still produced when no
Wave 2/3 data is on the model (the legacy "Sensitivity" sheet is
replaced by the W4.2 "Pricing Sensitivity" sheet — barebones deals
that lack sensitivity_grid drop the sensitivity sheet entirely rather
than synthesizing one).

Number formats follow institutional convention (currency, percent, x).
"""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ─────────────────────────── styling helpers ───────────────────────────

USD_FMT = '$#,##0;[Red]($#,##0)'
USD_2_FMT = '$#,##0.00'
PCT_FMT = '0.00%'
MULT_FMT = '0.00"x"'
INT_FMT = '#,##0'

HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # slate-800
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill("solid", fgColor="E5E7EB")  # gray-200
SUBHEAD_FONT = Font(bold=True, color="111827", size=10)
TOTAL_FILL = PatternFill("solid", fgColor="FEF3C7")  # amber-100
TOTAL_FONT = Font(bold=True, color="111827", size=10)
BRAND_FILL = PatternFill("solid", fgColor="F59E0B")  # amber-500
BRAND_FONT = Font(bold=True, color="FFFFFF", size=14)
LIGHT_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# Source-color map for op-ratio provenance rows. T-12 = green (best
# anchor), Portfolio = blue (in-house brand), CBRE = amber (market
# brand-light), HOST = gray (neutral fallback), Override = pink (warn —
# analyst intervened).
_SOURCE_ROW_FILL: dict[str, PatternFill] = {
    "t12_actual": PatternFill("solid", fgColor="DCFCE7"),       # green
    "portfolio_pnl": PatternFill("solid", fgColor="DBEAFE"),    # blue
    "cbre_horizons": PatternFill("solid", fgColor="FEF3C7"),    # amber
    "pnl_benchmark": PatternFill("solid", fgColor="E5E7EB"),    # gray
    "analyst_override": PatternFill("solid", fgColor="FCE7F3"), # pink
    "seed": PatternFill("solid", fgColor="F3F4F6"),             # neutral
}

_SOURCE_DISPLAY: dict[str, str] = {
    "t12_actual": "T-12",
    "portfolio_pnl": "Portfolio",
    "cbre_horizons": "CBRE",
    "pnl_benchmark": "HOST",
    "analyst_override": "Override",
    "seed": "Seed",
}


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = LIGHT_BORDER


def _style_subhead_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUBHEAD_FILL
        cell.font = SUBHEAD_FONT
        cell.border = LIGHT_BORDER


def _style_total_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = LIGHT_BORDER


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze_top(ws: Worksheet, row: int = 2) -> None:
    ws.freeze_panes = ws.cell(row=row, column=1)


def _apply_color_scale(
    ws: Worksheet,
    range_str: str,
    *,
    lo_color: str = "FECACA",  # red-200
    mid_color: str = "FEF3C7", # amber-100
    hi_color: str = "BBF7D0",  # green-200
) -> None:
    """Attach a 3-colour conditional-format rule to a range.

    Uses ``openpyxl.formatting.rule.ColorScaleRule`` so the workbook
    carries a real conditional-format rule (analysts can see the rule
    in Excel's "Manage Rules" dialog rather than just baked-in fill).
    """
    rule = ColorScaleRule(
        start_type="min", start_color=lo_color,
        mid_type="percentile", mid_value=50, mid_color=mid_color,
        end_type="max", end_color=hi_color,
    )
    ws.conditional_formatting.add(range_str, rule)


# ─────────────────────────── Wave 2/3 aggregator ───────────────────────


def _aggregate_wave2_3_for_excel(model: dict[str, Any]) -> dict[str, Any]:
    """Read Wave 2/3 engine outputs off the ``model`` dict into a clean
    shape for the Excel sheet builders. Mirrors the memo PDF's
    ``_aggregate_wave2_for_memo`` so a single deal payload feeds both
    artifacts.

    Returns a dict with one key per Wave 2/3 section. Each key holds
    either the section's data (dict / list) OR ``None`` when the
    underlying engine hasn't been run. Downstream sheet builders treat
    ``None`` as "skip the sheet entirely" — that's how a barebones
    deal still produces a valid model.
    """
    # ── Revenue segments (multi-year preferred) ──────────────────────
    segments_by_year = model.get("segments_by_year")
    rev = model.get("revenue_engine") or {}
    if not segments_by_year:
        segments_by_year = rev.get("segments_by_year") or rev.get("segment_breakdown")
    if not segments_by_year or not isinstance(segments_by_year, list):
        segments_by_year = None

    # ── PIP displacement ─────────────────────────────────────────────
    pip = model.get("pip_displacement")
    if pip and isinstance(pip, dict):
        if not pip.get("closure_strategy") or pip.get("closure_strategy") == "none":
            pip = None
    else:
        pip = None

    # ── Capex 3-bucket schedule ──────────────────────────────────────
    capex_schedule = model.get("capex_schedule")
    if not capex_schedule or not isinstance(capex_schedule, list):
        capex_schedule = None

    # ── Op-ratio provenance ─────────────────────────────────────────
    op_prov = model.get("op_ratio_provenance")
    if not op_prov or not isinstance(op_prov, dict) or not op_prov.get("lines"):
        op_prov = None

    # ── Pricing sensitivity grid ─────────────────────────────────────
    grid = model.get("sensitivity_grid")
    if not grid or not isinstance(grid, dict) or not grid.get("cells"):
        grid = None

    # ── Max-price result ─────────────────────────────────────────────
    max_price = model.get("max_price")
    if not max_price or not isinstance(max_price, dict):
        max_price = None

    # ── Historical baseline ─────────────────────────────────────────
    baseline = model.get("historical_baseline")
    if not baseline or not isinstance(baseline, dict) or not baseline.get("years"):
        baseline = None
    elif float(baseline.get("coverage_pct") or 0) <= 0:
        baseline = None

    # ── Comp Sales (W3.1) ────────────────────────────────────────────
    comp_sales = model.get("comp_sales")
    if not comp_sales or not isinstance(comp_sales, dict) or not comp_sales.get("transactions"):
        comp_sales = None

    # ── STR forward forecast (W3.3) ──────────────────────────────────
    str_forecast = model.get("str_forecast")
    if not str_forecast or not isinstance(str_forecast, dict):
        str_forecast = None
    else:
        hist = str_forecast.get("historical_months") or []
        fcst = str_forecast.get("forecast_months") or {}
        if not hist and not fcst:
            str_forecast = None

    # ── Named scenarios (W3.2) ──────────────────────────────────────
    named_scenarios = model.get("named_scenarios")
    if (
        not named_scenarios
        or not isinstance(named_scenarios, list)
        or len(named_scenarios) < 2
    ):
        # Need at least 2 scenarios for a compare table.
        named_scenarios = None

    # ── LOI draft ────────────────────────────────────────────────────
    loi = model.get("loi_draft")
    if not loi or not isinstance(loi, dict) or not loi.get("rendered_markdown"):
        loi = None

    return {
        "segments_by_year": segments_by_year,
        "pip": pip,
        "capex_schedule": capex_schedule,
        "op_ratio_provenance": op_prov,
        "sensitivity_grid": grid,
        "max_price": max_price,
        "historical_baseline": baseline,
        "comp_sales": comp_sales,
        "str_forecast": str_forecast,
        "named_scenarios": named_scenarios,
        "loi": loi,
    }


# ─────────────────────────── tab builders ───────────────────────────


def _build_cover(wb: Workbook, model: dict[str, Any], sections: list[str]) -> None:
    """Cover sheet — refreshed for W4.2.

    Adds: a "Quick Read" block of six headline KPIs, a "Sections
    Included" list that automatically reflects which Wave 2/3 sheets
    actually rendered, and an explicit "Generated by Fondok · v3.x ·
    UTC timestamp" stamp.

    Caller is expected to have reserved the active workbook sheet
    (``wb.active``) for the Cover; this function populates it.
    """
    ws = wb.active
    ws.title = "Cover"

    deal_name = model.get("deal_name", "Hotel Underwriting Model")
    location = model.get("location", "")
    keys = model.get("keys", "")
    brand = model.get("brand", "")

    inv = model.get("investment_engine", {})
    ret = model.get("returns_engine", {})
    debt = model.get("debt_engine", {})
    pl_lines = (model.get("p_and_l_engine_proforma") or {}).get("lines", [])
    noi_y1 = next(
        (row.get("y1", 0) * 1000 for row in pl_lines if row.get("label") == "Net Operating Income"),
        0,
    )
    noi_stab = next(
        (row.get("y5", 0) * 1000 for row in pl_lines if row.get("label") == "Net Operating Income"),
        0,
    )

    # Title ribbon
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "FONDOK ACQUISITION MODEL"
    title.fill = BRAND_FILL
    title.font = BRAND_FONT
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Property"
    ws["B3"] = deal_name
    ws["A4"] = "Location"
    ws["B4"] = location
    ws["A5"] = "Brand"
    ws["B5"] = brand
    ws["A6"] = "Keys"
    ws["B6"] = keys
    ws["A7"] = "Generated"
    ws["B7"] = (
        f"Generated by Fondok · "
        f"{model.get('model_version', 'fondok-engine-3.x')} · "
        f"{datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"
    )
    ws["A8"] = "Model Version"
    ws["B8"] = model.get("model_version", "fondok-engine-3.x")

    for r in range(3, 9):
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)

    # ── Quick Read — 6 headline numbers ─────────────────────────────
    ws.merge_cells("A10:F10")
    ws["A10"] = "QUICK READ"
    _style_header_row(ws, 10, 6)

    quick = [
        ("Purchase Price", inv.get("purchase_price_usd", 0), USD_FMT),
        ("$ / Key", inv.get("price_per_key_usd", 0), USD_FMT),
        ("Year 1 NOI", noi_y1, USD_FMT),
        ("Stabilized NOI", noi_stab, USD_FMT),
        ("Levered IRR", ret.get("levered_irr", 0), PCT_FMT),
        ("Equity Multiple", ret.get("equity_multiple", 0), MULT_FMT),
    ]
    row = 11
    for label, val, fmt in quick:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = fmt
        row += 1

    # ── Key Metrics block (existing) ────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="KEY METRICS")
    _style_header_row(ws, row, 6)
    row += 1

    metrics = [
        ("Total Capital", inv.get("total_capital_usd", 0), USD_FMT),
        ("Loan Amount", debt.get("loan_amount_usd", 0), USD_FMT),
        ("Loan Rate", debt.get("interest_rate_pct", 0), PCT_FMT),
        ("Year 1 DSCR", debt.get("year1_dscr", 0), MULT_FMT),
        ("Year 1 CoC", ret.get("year1_cash_on_cash", 0), PCT_FMT),
        ("Hold (yrs)", ret.get("hold_years", 0), INT_FMT),
        ("Exit Cap Rate", ret.get("exit_cap_rate_pct", 0), PCT_FMT),
        ("Entry Cap (Y1 UW)", inv.get("entry_cap_rate_year1_uw", 0), PCT_FMT),
    ]
    for label, val, fmt in metrics:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = fmt
        row += 1

    # ── Sections Included — conditional list ────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="SECTIONS INCLUDED")
    _style_header_row(ws, row, 6)
    row += 1
    for name in sections:
        ws.cell(row=row, column=1, value=name).font = Font(size=10)
        row += 1

    _autosize(ws, [28, 30, 4, 22, 22, 4])
    _freeze_top(ws, row=2)


def _build_assumptions(wb: Workbook, model: dict[str, Any]) -> None:
    ws = wb.create_sheet("Assumptions")
    inv = model.get("investment_engine", {})
    debt = model.get("debt_engine", {})
    refi = model.get("refi_engine", {})
    ret = model.get("returns_engine", {})

    ws["A1"] = "Assumption"
    ws["B1"] = "Value"
    ws["C1"] = "Unit / Notes"
    _style_header_row(ws, 1, 3)

    sections: list[tuple[str, list[tuple[str, Any, str, str]]]] = [
        (
            "Acquisition",
            [
                ("Purchase Price", inv.get("purchase_price_usd", 0), USD_FMT, "USD"),
                ("Price / Key", inv.get("price_per_key_usd", 0), USD_FMT, "USD/key"),
                ("Closing Costs", inv.get("closing_costs_usd", 0), USD_FMT, "USD"),
                ("Renovation Budget", inv.get("renovation_budget_usd", 0), USD_FMT, "PIP / capex"),
                ("Renovation / Key", inv.get("renovation_per_key_usd", 0), USD_FMT, "USD/key"),
                ("Soft Costs", inv.get("soft_costs_usd", 0), USD_FMT, ""),
                ("Contingency", inv.get("contingency_usd", 0), USD_FMT, ""),
                ("Working Capital", inv.get("working_capital_usd", 0), USD_FMT, ""),
                ("Loan Costs", inv.get("loan_costs_usd", 0), USD_FMT, ""),
                ("Total Capital", inv.get("total_capital_usd", 0), USD_FMT, "All-in"),
                ("Total Capital / Key", inv.get("total_capital_per_key_usd", 0), USD_FMT, ""),
                ("Entry Cap (T-12)", inv.get("entry_cap_rate_t12", 0), PCT_FMT, ""),
                ("Entry Cap (Y1 UW)", inv.get("entry_cap_rate_year1_uw", 0), PCT_FMT, ""),
                ("Year 1 Yield on Cost", inv.get("year1_yield_on_cost", 0), PCT_FMT, ""),
            ],
        ),
        (
            "Senior Debt",
            [
                ("Loan Amount", debt.get("loan_amount_usd", 0), USD_FMT, ""),
                ("LTV (cost)", debt.get("ltv_cost", 0), PCT_FMT, ""),
                ("LTV (value)", debt.get("ltv_value", 0), PCT_FMT, ""),
                ("LTC", debt.get("ltc", 0), PCT_FMT, ""),
                ("Interest Rate", debt.get("interest_rate_pct", 0), PCT_FMT, ""),
                ("Amortization (yrs)", debt.get("amortization_years", 0), INT_FMT, ""),
                ("Term (yrs)", debt.get("term_years", 0), INT_FMT, ""),
                ("Annual Debt Service", debt.get("annual_debt_service_usd", 0), USD_FMT, ""),
                ("Year 1 DSCR", debt.get("year1_dscr", 0), MULT_FMT, ""),
                ("Year 1 Debt Yield", debt.get("year1_debt_yield", 0), PCT_FMT, ""),
                ("IO Period (months)", debt.get("interest_only_period_months", 0), INT_FMT, ""),
            ],
        ),
        (
            "Refinance",
            [
                ("Refi Year", refi.get("refi_year", 0), INT_FMT, ""),
                ("Refi LTV", refi.get("refi_ltv", 0), PCT_FMT, ""),
                ("Refi Rate", refi.get("refi_rate_pct", 0), PCT_FMT, ""),
                ("Refi Term (yrs)", refi.get("refi_term_years", 0), INT_FMT, ""),
                ("Refi Amort (yrs)", refi.get("refi_amortization_years", 0), INT_FMT, ""),
                ("Refi Proceeds", refi.get("refi_proceeds_usd", 0), USD_FMT, ""),
                ("Cash Out", refi.get("cash_out_to_equity_usd", 0), USD_FMT, ""),
            ],
        ),
        (
            "Reversion / Returns",
            [
                ("Hold (yrs)", ret.get("hold_years", 0), INT_FMT, ""),
                ("Exit Cap Rate", ret.get("exit_cap_rate_pct", 0), PCT_FMT, ""),
                ("Terminal NOI", ret.get("terminal_noi_usd", 0), USD_FMT, ""),
                ("Gross Sale Price", ret.get("gross_sale_price_usd", 0), USD_FMT, ""),
                ("Selling Costs", ret.get("selling_costs_usd", 0), USD_FMT, ""),
                ("Net Sale Proceeds", ret.get("net_sale_proceeds_usd", 0), USD_FMT, ""),
            ],
        ),
    ]

    row = 2
    for section_name, items in sections:
        ws.cell(row=row, column=1, value=section_name)
        _style_subhead_row(ws, row, 3)
        row += 1
        for label, val, fmt, note in items:
            ws.cell(row=row, column=1, value=label)
            c = ws.cell(row=row, column=2, value=val)
            c.number_format = fmt
            ws.cell(row=row, column=3, value=note)
            row += 1
        row += 1

    _autosize(ws, [32, 22, 30])
    _freeze_top(ws, row=2)


def _build_sources_uses(wb: Workbook, model: dict[str, Any]) -> None:
    ws = wb.create_sheet("Sources & Uses")
    inv = model.get("investment_engine", {})
    debt = model.get("debt_engine", {})
    sources = model.get("sources", []) or [
        {
            "label": "Senior Debt",
            "amount": debt.get("loan_amount_usd", 0),
            "pct": debt.get("ltc", 0),
        },
        {
            "label": "Equity",
            "amount": inv.get("total_capital_usd", 0) - debt.get("loan_amount_usd", 0),
            "pct": 1 - debt.get("ltc", 0),
        },
    ]
    uses = model.get("uses", []) or [
        {"label": "Purchase Price", "amount": inv.get("purchase_price_usd", 0)},
        {"label": "Closing Costs", "amount": inv.get("closing_costs_usd", 0)},
        {"label": "Renovation", "amount": inv.get("renovation_budget_usd", 0)},
        {"label": "Soft Costs", "amount": inv.get("soft_costs_usd", 0)},
        {"label": "Contingency", "amount": inv.get("contingency_usd", 0)},
        {"label": "Working Capital", "amount": inv.get("working_capital_usd", 0)},
        {"label": "Loan Costs", "amount": inv.get("loan_costs_usd", 0)},
    ]

    ws["A1"] = "SOURCES"
    ws["B1"] = "Amount"
    ws["C1"] = "% of Total"
    _style_header_row(ws, 1, 3)
    row = 2
    src_total = sum(s.get("amount", 0) for s in sources if not s.get("total"))
    for s in sources:
        if s.get("total"):
            continue
        ws.cell(row=row, column=1, value=s["label"])
        c = ws.cell(row=row, column=2, value=s["amount"])
        c.number_format = USD_FMT
        pct = s.get("pct") if s.get("pct") is not None else (s["amount"] / src_total if src_total else 0)
        p = ws.cell(row=row, column=3, value=pct)
        p.number_format = PCT_FMT
        row += 1
    ws.cell(row=row, column=1, value="Total Sources")
    c = ws.cell(row=row, column=2, value=f"=SUM(B2:B{row - 1})")
    c.number_format = USD_FMT
    ws.cell(row=row, column=3, value=1.0).number_format = PCT_FMT
    _style_total_row(ws, row, 3)
    row += 3

    ws.cell(row=row, column=1, value="USES")
    ws.cell(row=row, column=2, value="Amount")
    ws.cell(row=row, column=3, value="% of Total")
    _style_header_row(ws, row, 3)
    use_start = row + 1
    use_total = sum(u.get("amount", 0) for u in uses if not u.get("total"))
    row = use_start
    for u in uses:
        if u.get("total"):
            continue
        ws.cell(row=row, column=1, value=u["label"])
        c = ws.cell(row=row, column=2, value=u["amount"])
        c.number_format = USD_FMT
        pct = (u["amount"] / use_total) if use_total else 0
        p = ws.cell(row=row, column=3, value=pct)
        p.number_format = PCT_FMT
        row += 1
    ws.cell(row=row, column=1, value="Total Uses")
    c = ws.cell(row=row, column=2, value=f"=SUM(B{use_start}:B{row - 1})")
    c.number_format = USD_FMT
    ws.cell(row=row, column=3, value=1.0).number_format = PCT_FMT
    _style_total_row(ws, row, 3)

    _autosize(ws, [28, 22, 14])
    _freeze_top(ws, row=2)


def _build_proforma(wb: Workbook, model: dict[str, Any]) -> None:
    ws = wb.create_sheet("Operating Proforma")
    pl = model.get("p_and_l_engine_proforma", {})
    lines = pl.get("lines", [])

    headers = ["Line Item", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "CAGR"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    for r, line in enumerate(lines, start=2):
        ws.cell(row=r, column=1, value=line["label"])
        for i, key in enumerate(["y1", "y2", "y3", "y4", "y5"], start=2):
            c = ws.cell(row=r, column=i, value=line.get(key, 0) * 1000)
            c.number_format = USD_FMT
        if line.get("cagr") is not None:
            c = ws.cell(row=r, column=7, value=line["cagr"])
            c.number_format = PCT_FMT
        if line.get("bold"):
            _style_total_row(ws, r, len(headers))

    period_row = len(lines) + 3
    ws.cell(row=period_row, column=1, value="Period:").font = Font(italic=True)
    ws.cell(row=period_row, column=2, value=pl.get("year1_period", "")).font = Font(italic=True)
    ws.cell(row=period_row + 1, column=1, value="Note: figures in USD (whole dollars).").font = Font(italic=True, color="6B7280")

    _autosize(ws, [30, 16, 16, 16, 16, 16, 12])
    _freeze_top(ws, row=2)


def _build_debt_schedule(wb: Workbook, model: dict[str, Any]) -> None:
    ws = wb.create_sheet("Debt Schedule")
    debt = model.get("debt_engine", {})

    loan = float(debt.get("loan_amount_usd", 0))
    rate = float(debt.get("interest_rate_pct", 0))
    amort_years = int(debt.get("amortization_years", 30))
    term_years = int(debt.get("term_years", 5))
    io_months = int(debt.get("interest_only_period_months", 0))

    monthly_rate = rate / 12 if rate else 0
    n = amort_years * 12
    if monthly_rate > 0:
        pmt = loan * (monthly_rate * (1 + monthly_rate) ** n) / ((1 + monthly_rate) ** n - 1)
    else:
        pmt = loan / n if n else 0

    headers = ["Month", "Beg Balance", "Payment", "Interest", "Principal", "End Balance"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    bal = loan
    months = term_years * 12
    for m in range(1, months + 1):
        beg = bal
        interest = beg * monthly_rate
        if m <= io_months:
            principal = 0.0
            payment = interest
        else:
            payment = pmt
            principal = payment - interest
        end = beg - principal

        row = m + 1
        ws.cell(row=row, column=1, value=m).number_format = INT_FMT
        c = ws.cell(row=row, column=2, value=beg)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=3, value=payment)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=4, value=interest)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=5, value=principal)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=6, value=end)
        c.number_format = USD_FMT

        bal = end

    total_row = months + 2
    ws.cell(row=total_row, column=1, value="Totals").font = Font(bold=True)
    c = ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{months + 1})")
    c.number_format = USD_FMT
    c = ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{months + 1})")
    c.number_format = USD_FMT
    c = ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{months + 1})")
    c.number_format = USD_FMT
    _style_total_row(ws, total_row, len(headers))

    _autosize(ws, [8, 18, 16, 16, 16, 18])
    _freeze_top(ws, row=2)


def _build_returns(wb: Workbook, model: dict[str, Any]) -> None:
    ws = wb.create_sheet("Returns")
    ret = model.get("returns_engine", {})
    cf = model.get("cash_flow_engine", {})

    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    _style_header_row(ws, 1, 2)

    items = [
        ("Hold Period (yrs)", ret.get("hold_years", 0), INT_FMT),
        ("Exit Cap Rate", ret.get("exit_cap_rate_pct", 0), PCT_FMT),
        ("Terminal NOI", ret.get("terminal_noi_usd", 0), USD_FMT),
        ("Gross Sale Price", ret.get("gross_sale_price_usd", 0), USD_FMT),
        ("Selling Costs", ret.get("selling_costs_usd", 0), USD_FMT),
        ("Net Sale Proceeds", ret.get("net_sale_proceeds_usd", 0), USD_FMT),
        ("Levered IRR", ret.get("levered_irr", 0), PCT_FMT),
        ("Unlevered IRR", ret.get("unlevered_irr", 0), PCT_FMT),
        ("Equity Multiple", ret.get("equity_multiple", 0), MULT_FMT),
        ("Year 1 Cash-on-Cash", ret.get("year1_cash_on_cash", 0), PCT_FMT),
        ("Average Cash-on-Cash", ret.get("avg_cash_on_cash", 0), PCT_FMT),
    ]
    row = 2
    for label, val, fmt in items:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = fmt
        row += 1

    debt = model.get("debt_engine", {})
    ws.cell(row=row, column=1, value="Year 1 DSCR")
    c = ws.cell(row=row, column=2, value=debt.get("year1_dscr", 0))
    c.number_format = MULT_FMT
    row += 1
    ws.cell(row=row, column=1, value="Year 1 Debt Yield")
    c = ws.cell(row=row, column=2, value=debt.get("year1_debt_yield", 0))
    c.number_format = PCT_FMT
    row += 2

    ws.cell(row=row, column=1, value="CASH FLOW WATERFALL ($)")
    _style_subhead_row(ws, row, 6)
    row += 1
    ws.cell(row=row, column=1, value="Year")
    for i in range(1, 6):
        ws.cell(row=row, column=i + 1, value=f"Y{i}")
    _style_header_row(ws, row, 6)
    row += 1

    ws.cell(row=row, column=1, value="Cash Flow After Debt")
    for i, key in enumerate(
        ["year1_cf_after_debt_usd", "year2_cf_after_debt_usd", "year3_cf_after_debt_usd",
         "year4_cf_after_debt_usd", "year5_cf_after_debt_usd"],
        start=2,
    ):
        c = ws.cell(row=row, column=i, value=cf.get(key, 0))
        c.number_format = USD_FMT
    row += 1

    ws.cell(row=row, column=1, value="Cumulative")
    cumulative = 0
    for i in range(2, 7):
        cumulative += cf.get(f"year{i - 1}_cf_after_debt_usd", 0)
        c = ws.cell(row=row, column=i, value=cumulative)
        c.number_format = USD_FMT
    _style_total_row(ws, row, 6)

    _autosize(ws, [28, 18, 18, 18, 18, 18])
    _freeze_top(ws, row=2)


def _build_partnership(wb: Workbook, model: dict[str, Any]) -> None:
    ws = wb.create_sheet("Partnership")
    p = model.get("partnership_engine", {})

    ws["A1"] = "Partnership Structure"
    _style_header_row(ws, 1, 3)
    rows = [
        ("Structure", p.get("structure", "GP/LP waterfall"), ""),
        ("LP Equity", p.get("lp_equity_usd", 0), USD_FMT),
        ("GP Equity", p.get("gp_equity_usd", 0), USD_FMT),
        ("Total Equity", p.get("total_equity_usd", 0), USD_FMT),
        ("LP Preferred Return", p.get("lp_pref_pct", 0), PCT_FMT),
    ]
    row = 2
    for label, val, fmt in rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=val)
        if fmt:
            c.number_format = fmt
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Waterfall Tiers")
    _style_header_row(ws, row, 3)
    row += 1
    ws.cell(row=row, column=1, value="Tier")
    ws.cell(row=row, column=2, value="IRR Hurdle")
    ws.cell(row=row, column=3, value="GP Promote")
    _style_subhead_row(ws, row, 3)
    row += 1
    tier_rows = [
        ("Tier 1 — Pref + Promote", p.get("gp_promote_tier_1_irr_hurdle", 0), p.get("gp_promote_tier_1_pct", 0)),
        ("Tier 2 — Catch-up + Promote", p.get("gp_promote_tier_2_irr_hurdle", 0), p.get("gp_promote_tier_2_pct", 0)),
    ]
    for label, hurdle, promote in tier_rows:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=hurdle)
        c.number_format = PCT_FMT
        c = ws.cell(row=row, column=3, value=promote)
        c.number_format = PCT_FMT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Distributions Summary")
    _style_header_row(ws, row, 3)
    row += 1
    summary = [
        ("LP IRR (after promote)", p.get("lp_irr_after_promote", 0), PCT_FMT),
        ("GP IRR (after promote)", p.get("gp_irr_after_promote", 0), PCT_FMT),
        ("LP Equity Multiple", p.get("lp_equity_multiple", 0), MULT_FMT),
        ("GP Equity Multiple", p.get("gp_equity_multiple", 0), MULT_FMT),
    ]
    for label, val, fmt in summary:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = fmt
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Distribution Per Year (estimate)")
    _style_header_row(ws, row, 6)
    row += 1
    headers = ["Year", "Total CF", "LP Share", "GP Share", "LP %", "GP %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_subhead_row(ws, row, len(headers))
    row += 1

    cf = model.get("cash_flow_engine", {})
    lp_eq = float(p.get("lp_equity_usd", 0))
    gp_eq = float(p.get("gp_equity_usd", 0))
    total_eq = lp_eq + gp_eq if (lp_eq + gp_eq) > 0 else 1
    lp_pct = lp_eq / total_eq
    gp_pct = gp_eq / total_eq
    for y in range(1, 6):
        total = cf.get(f"year{y}_cf_after_debt_usd", 0)
        ws.cell(row=row, column=1, value=f"Y{y}")
        c = ws.cell(row=row, column=2, value=total)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=3, value=total * lp_pct)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=4, value=total * gp_pct)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=5, value=lp_pct)
        c.number_format = PCT_FMT
        c = ws.cell(row=row, column=6, value=gp_pct)
        c.number_format = PCT_FMT
        row += 1

    _autosize(ws, [30, 18, 18, 18, 12, 12])
    _freeze_top(ws, row=2)


def _build_variance(wb: Workbook, model: dict[str, Any]) -> None:
    ws = wb.create_sheet("Variance")
    flags = model.get("variance_flags", []) or []

    headers = ["Flag ID", "Severity", "Metric", "Broker / Value", "T-12 / Threshold", "Variance", "Action"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    severity_fills = {
        "CRITICAL": PatternFill("solid", fgColor="FEE2E2"),
        "WARN": PatternFill("solid", fgColor="FEF3C7"),
        "INFO": PatternFill("solid", fgColor="DBEAFE"),
    }

    for r, f in enumerate(flags, start=2):
        ws.cell(row=r, column=1, value=f.get("flag_id", ""))
        sev = f.get("severity", "INFO")
        sev_cell = ws.cell(row=r, column=2, value=sev)
        sev_cell.fill = severity_fills.get(sev, severity_fills["INFO"])
        sev_cell.font = Font(bold=True)
        ws.cell(row=r, column=3, value=f.get("metric", ""))
        if "broker_value" in f:
            ws.cell(row=r, column=4, value=f["broker_value"])
        elif "value" in f:
            ws.cell(row=r, column=4, value=f["value"])
        if "t12_value" in f:
            ws.cell(row=r, column=5, value=f["t12_value"])
        elif "threshold_max" in f:
            ws.cell(row=r, column=5, value=f"{f.get('threshold_min','')}–{f.get('threshold_max','')}")
        if "variance_pct" in f:
            v = ws.cell(row=r, column=6, value=f["variance_pct"])
            v.number_format = PCT_FMT
        elif "variance_pct_pts" in f:
            ws.cell(row=r, column=6, value=f"{f['variance_pct_pts']} pts")
        ws.cell(row=r, column=7, value=f.get("recommended_action", ""))
        ws.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical="top")

    _autosize(ws, [10, 12, 22, 18, 18, 14, 60])
    _freeze_top(ws, row=2)


def _build_market_comps(wb: Workbook, model: dict[str, Any]) -> None:
    ws = wb.create_sheet("Market Comps")
    comps = model.get("market_comps", []) or []

    headers = ["Property", "Keys", "Date", "Price", "Per Key", "Cap Rate", "Buyer"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    for r, c in enumerate(comps, start=2):
        ws.cell(row=r, column=1, value=c.get("name", ""))
        ws.cell(row=r, column=2, value=c.get("keys", 0)).number_format = INT_FMT
        ws.cell(row=r, column=3, value=c.get("date", ""))
        ws.cell(row=r, column=4, value=c.get("price", ""))
        ws.cell(row=r, column=5, value=c.get("per_key", c.get("perKey", "")))
        ws.cell(row=r, column=6, value=c.get("cap", ""))
        ws.cell(row=r, column=7, value=c.get("buyer", ""))

    _autosize(ws, [32, 8, 14, 14, 14, 12, 26])
    _freeze_top(ws, row=2)


# ──────────────────────── Wave 2/3 sheet builders ──────────────────────


def _build_revenue_mix(wb: Workbook, segments_by_year: list[dict[str, Any]] | None) -> bool:
    """Segment-by-year revenue mix table (5 segments × hold years).

    Cols per segment: Mix %, ADR, Gross Rev, Channel Cost %, Net Rev.
    Includes a Total row and an OTA-flag highlight when any segment
    carries ≥15% channel cost. Returns True iff the sheet rendered.
    """
    if not segments_by_year:
        return False

    ws = wb.create_sheet("Revenue Mix")

    headers = [
        "Segment", "Year", "Mix %", "ADR", "Gross Revenue",
        "Channel Cost %", "Net Revenue",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    ota_fill = PatternFill("solid", fgColor="FEF3C7")  # amber for OTA-heavy
    row = 2
    total_gross = 0.0
    total_net = 0.0
    for year_block in segments_by_year:
        year = year_block.get("year")
        breakdown = (
            year_block.get("segment_breakdown")
            or year_block.get("segments")
            or []
        )
        # Subhead for this year
        ws.cell(row=row, column=1, value=f"Year {year}")
        _style_subhead_row(ws, row, len(headers))
        row += 1
        year_gross = 0.0
        year_net = 0.0
        for seg in breakdown:
            ws.cell(row=row, column=1, value=str(seg.get("name", "—")))
            ws.cell(row=row, column=2, value=year).number_format = INT_FMT
            c = ws.cell(row=row, column=3, value=float(seg.get("mix_pct") or 0))
            c.number_format = PCT_FMT
            c = ws.cell(row=row, column=4, value=float(seg.get("adr") or 0))
            c.number_format = USD_FMT
            gross_rev = float(seg.get("gross_revenue") or 0)
            c = ws.cell(row=row, column=5, value=gross_rev)
            c.number_format = USD_FMT
            channel = float(seg.get("channel_cost_pct") or 0)
            c = ws.cell(row=row, column=6, value=channel)
            c.number_format = PCT_FMT
            net_rev = float(seg.get("net_revenue") or 0)
            c = ws.cell(row=row, column=7, value=net_rev)
            c.number_format = USD_FMT
            # Highlight OTA-heavy segments
            if channel >= 0.15:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = ota_fill
            year_gross += gross_rev
            year_net += net_rev
            row += 1
        # Year total
        ws.cell(row=row, column=1, value=f"Total Y{year}")
        c = ws.cell(row=row, column=5, value=year_gross)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=7, value=year_net)
        c.number_format = USD_FMT
        _style_total_row(ws, row, len(headers))
        total_gross += year_gross
        total_net += year_net
        row += 1

    # Grand total across all years
    ws.cell(row=row, column=1, value="GRAND TOTAL")
    c = ws.cell(row=row, column=5, value=total_gross)
    c.number_format = USD_FMT
    c = ws.cell(row=row, column=7, value=total_net)
    c.number_format = USD_FMT
    _style_total_row(ws, row, len(headers))

    _autosize(ws, [24, 8, 12, 14, 18, 16, 18])
    _freeze_top(ws, row=2)
    return True


def _build_renovation_plan(wb: Workbook, pip: dict[str, Any] | None) -> bool:
    """PIP displacement summary sheet: closure strategy, schedule, Y1
    displacement $, Y2 recovery curve, brand recovery multipliers.
    """
    if not pip:
        return False

    ws = wb.create_sheet("Renovation Plan")

    strategy_label = {
        "rolling": "Rolling",
        "full_closure": "Full Closure",
        "wing_by_wing": "Wing-by-Wing",
    }.get(pip.get("closure_strategy") or "", str(pip.get("closure_strategy") or "—"))

    ws["A1"] = "Renovation Plan (PIP Displacement)"
    _style_header_row(ws, 1, 4)

    row = 2
    summary = [
        ("Closure Strategy", strategy_label, ""),
        ("Brand", pip.get("brand") or "Independent", ""),
        ("Recovery Months", pip.get("occupancy_recovery_months") or 0, INT_FMT),
        ("RevPAR Index (post-reno)", pip.get("revpar_index_post_reno") or 1.0, MULT_FMT),
        ("Y1 Displacement", pip.get("y1_displacement_usd") or 0, USD_FMT),
    ]
    for label, val, fmt in summary:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=val)
        if fmt:
            c.number_format = fmt
        row += 1

    # Brand recovery multipliers (informational note + table)
    row += 1
    ws.cell(row=row, column=1, value="BRAND RECOVERY MULTIPLIERS")
    _style_subhead_row(ws, row, 4)
    row += 1
    brand_label = (pip.get("brand") or "Independent").title()
    ws.cell(row=row, column=1, value=f"Selected brand: {brand_label}")
    ws.cell(row=row, column=1).font = Font(italic=True)
    row += 1
    ws.cell(row=row, column=1, value="Brand")
    ws.cell(row=row, column=2, value="Post-Reno RevPAR Index")
    ws.cell(row=row, column=3, value="Typical Recovery (months)")
    _style_subhead_row(ws, row, 4)
    row += 1
    for brand_name, idx, mo in (
        ("Marriott Full Service", 1.12, 4),
        ("Hilton Full Service", 1.10, 5),
        ("Hyatt Full Service", 1.10, 5),
        ("Kimpton / Lifestyle", 1.08, 6),
        ("Independent", 1.05, 8),
    ):
        ws.cell(row=row, column=1, value=brand_name)
        c = ws.cell(row=row, column=2, value=idx)
        c.number_format = MULT_FMT
        c = ws.cell(row=row, column=3, value=mo)
        c.number_format = INT_FMT
        row += 1

    # Month-by-month offline schedule
    row += 1
    ws.cell(row=row, column=1, value="MONTHLY OFFLINE SCHEDULE (Y1)")
    _style_subhead_row(ws, row, 4)
    row += 1
    sched = pip.get("pct_rooms_offline_by_month") or []
    ws.cell(row=row, column=1, value="Month")
    ws.cell(row=row, column=2, value="% Rooms Offline")
    _style_header_row(ws, row, 2)
    row += 1
    for m, pct in enumerate(sched, start=1):
        ws.cell(row=row, column=1, value=f"M{m}").number_format = INT_FMT
        c = ws.cell(row=row, column=2, value=float(pct))
        c.number_format = PCT_FMT
        row += 1

    # Y2 recovery curve
    row += 1
    ws.cell(row=row, column=1, value="Y2 RECOVERY CURVE")
    _style_subhead_row(ws, row, 4)
    row += 1
    ws.cell(row=row, column=1, value="Month")
    ws.cell(row=row, column=2, value="% of Baseline RevPAR")
    _style_header_row(ws, row, 2)
    row += 1
    for i, v in enumerate(pip.get("y2_recovery_curve") or [], start=1):
        ws.cell(row=row, column=1, value=f"M{i}").number_format = INT_FMT
        c = ws.cell(row=row, column=2, value=float(v))
        c.number_format = PCT_FMT
        row += 1

    _autosize(ws, [32, 22, 22, 18])
    _freeze_top(ws, row=2)
    return True


def _build_capital_plan(wb: Workbook, schedule: list[dict[str, Any]] | None) -> bool:
    """Capital Plan — three-bucket capex (PIP / Non-PIP FF&E / ROI).

    Replaces the legacy single capex line in Assumptions. Rows are hold
    years plus a total row. Cell-level provenance comments distinguish
    analyst-override entries from defaults.
    """
    if not schedule:
        return False

    ws = wb.create_sheet("Capital Plan")
    headers = [
        "Year", "PIP", "Non-PIP FF&E", "ROI Investment",
        "ROI NOI Lift", "Total Capex",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    tot_pip = tot_non = tot_roi_inv = tot_roi_lift = tot_total = 0.0
    row = 2
    for y in schedule:
        year_label = y.get("year", "?")
        pip_v = float(y.get("pip_usd") or 0)
        non_v = float(y.get("non_pip_usd") or 0)
        roi_inv = float(y.get("roi_investment_usd") or 0)
        roi_lift = float(y.get("roi_noi_lift_usd") or 0)
        total = float(y.get("total_capex_usd") or pip_v + non_v + roi_inv)
        tot_pip += pip_v
        tot_non += non_v
        tot_roi_inv += roi_inv
        tot_roi_lift += roi_lift
        tot_total += total

        ws.cell(row=row, column=1, value=f"Year {year_label}")
        c = ws.cell(row=row, column=2, value=pip_v)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=3, value=non_v)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=4, value=roi_inv)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=5, value=roi_lift)
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=6, value=total)
        c.number_format = USD_FMT

        # Cell-level provenance note via openpyxl Comment, if set on
        # the schedule row.
        note = y.get("override_note")
        if note:
            ws.cell(row=row, column=2).comment = Comment(
                f"Override: {note}", "Fondok",
            )
        row += 1

    ws.cell(row=row, column=1, value="Total")
    c = ws.cell(row=row, column=2, value=tot_pip)
    c.number_format = USD_FMT
    c = ws.cell(row=row, column=3, value=tot_non)
    c.number_format = USD_FMT
    c = ws.cell(row=row, column=4, value=tot_roi_inv)
    c.number_format = USD_FMT
    c = ws.cell(row=row, column=5, value=tot_roi_lift)
    c.number_format = USD_FMT
    c = ws.cell(row=row, column=6, value=tot_total)
    c.number_format = USD_FMT
    _style_total_row(ws, row, len(headers))

    _autosize(ws, [12, 16, 18, 18, 18, 18])
    _freeze_top(ws, row=2)
    return True


def _build_op_ratio_provenance(wb: Workbook, op_prov: dict[str, Any] | None) -> bool:
    """Op-ratio provenance sheet — one row per ratio with winning source.

    Cols: Field, Value, Source, Why. Each row is tinted by source
    (T-12 = green, Portfolio = blue, CBRE = amber, HOST = neutral,
    Override = pink).
    """
    if not op_prov:
        return False
    lines = op_prov.get("lines") or []
    if not lines:
        return False

    ws = wb.create_sheet("Op-Ratio Provenance")
    headers = ["Field", "Value", "Source", "Document / Note"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    row = 2
    for line in lines:
        field = str(line.get("field") or "—")
        value = line.get("value")
        source = str(line.get("source") or "seed")
        doc = line.get("document_id") or "—"

        ws.cell(row=row, column=1, value=field)
        c = ws.cell(row=row, column=2, value=value if value is not None else 0)
        # Value-formatting heuristic — fractions get pct, big numbers get $.
        if isinstance(value, float) and 0 <= value <= 1:
            c.number_format = PCT_FMT
        else:
            c.number_format = USD_FMT
        ws.cell(row=row, column=3, value=_SOURCE_DISPLAY.get(source, source.title()))
        ws.cell(row=row, column=4, value=str(doc))

        # Row tint by source
        fill = _SOURCE_ROW_FILL.get(source)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
        row += 1

    # Legend
    row += 1
    ws.cell(row=row, column=1, value="Source precedence:").font = Font(bold=True, italic=True)
    ws.cell(row=row, column=2, value="Override → T-12 actuals → Portfolio → CBRE → HOST → Seed")
    ws.cell(row=row, column=2).font = Font(italic=True)

    _autosize(ws, [30, 16, 18, 40])
    _freeze_top(ws, row=2)
    return True


def _build_sensitivity_grid(wb: Workbook, grid: dict[str, Any] | None) -> bool:
    """Pricing Sensitivity — 5×5 IRR grid + 5×5 EM grid + max-price summary.

    Replaces the legacy synthesised Sensitivity sheet. IRR grid carries
    a 3-color scale conditional format; DSCR-breach cells get a red
    fill + ``!`` cell comment + red border. EM grid gets its own color
    scale.
    """
    if not grid:
        return False
    cells = grid.get("cells") or []
    if not cells:
        return False

    ws = wb.create_sheet("Pricing Sensitivity")
    target_irr = float(grid.get("target_irr") or 0.15)

    cap_axis = sorted({round(float(c.get("exit_cap_pct") or 0), 6) for c in cells})
    noi_axis = sorted(
        {round(float(c.get("noi_multiplier") or 0), 4) for c in cells},
        reverse=True,
    )
    cell_lookup = {
        (
            round(float(c.get("exit_cap_pct") or 0), 6),
            round(float(c.get("noi_multiplier") or 0), 4),
        ): c
        for c in cells
    }

    # ── Max-Price summary at top ────────────────────────────────────
    ws["A1"] = "MAX-PRICE FINDINGS"
    _style_header_row(ws, 1, len(cap_axis) + 2)
    row = 2
    max_price = grid.get("max_price") if isinstance(grid.get("max_price"), dict) else None
    ws.cell(row=row, column=1, value="Breakeven Exit Cap").font = Font(bold=True)
    bc = ws.cell(row=row, column=2, value=grid.get("breakeven_exit_cap_pct") or 0)
    bc.number_format = PCT_FMT
    ws.cell(row=row, column=4, value="Breakeven NOI Multiplier").font = Font(bold=True)
    bn = ws.cell(row=row, column=5, value=grid.get("breakeven_noi_multiplier") or 0)
    bn.number_format = MULT_FMT
    row += 1
    if max_price:
        ws.cell(row=row, column=1, value="Max Price for IRR target").font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=float(max_price.get("max_price_for_irr") or 0))
        c.number_format = USD_FMT
        ws.cell(row=row, column=4, value="Max Price for EM target").font = Font(bold=True)
        c = ws.cell(row=row, column=5, value=float(max_price.get("max_price_for_em") or 0))
        c.number_format = USD_FMT
        row += 1
        ws.cell(row=row, column=1, value="Binding Constraint").font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(max_price.get("binding_constraint") or "—").upper())
        row += 1
    row += 1

    # ── IRR grid header ─────────────────────────────────────────────
    ws.cell(row=row, column=1, value=f"LEVERED IRR (target = {target_irr:.2%})")
    _style_header_row(ws, row, len(cap_axis) + 2)
    row += 1
    ws.cell(row=row, column=1, value="NOI Multiplier \\ Exit Cap")
    for j, cap in enumerate(cap_axis, start=2):
        c = ws.cell(row=row, column=j, value=cap)
        c.number_format = PCT_FMT
        c.font = Font(bold=True)
    _style_subhead_row(ws, row, len(cap_axis) + 1)
    row += 1
    irr_start_row = row
    dscr_red_fill = PatternFill("solid", fgColor="FECACA")
    dscr_red_border = Border(
        left=Side(style="medium", color="DC2626"),
        right=Side(style="medium", color="DC2626"),
        top=Side(style="medium", color="DC2626"),
        bottom=Side(style="medium", color="DC2626"),
    )
    for nm in noi_axis:
        c = ws.cell(row=row, column=1, value=nm)
        c.number_format = MULT_FMT
        c.font = Font(bold=True)
        for j, cap in enumerate(cap_axis, start=2):
            cell_data = cell_lookup.get((cap, nm))
            if cell_data is None:
                ws.cell(row=row, column=j, value=None)
                continue
            irr = float(cell_data.get("levered_irr") or 0)
            cell = ws.cell(row=row, column=j, value=irr)
            cell.number_format = PCT_FMT
            if cell_data.get("breaches_dscr_floor"):
                cell.fill = dscr_red_fill
                cell.border = dscr_red_border
                cell.comment = Comment("! DSCR breach", "Fondok")
        row += 1
    irr_end_row = row - 1
    irr_first_col_letter = get_column_letter(2)
    irr_last_col_letter = get_column_letter(1 + len(cap_axis))
    irr_range = f"{irr_first_col_letter}{irr_start_row}:{irr_last_col_letter}{irr_end_row}"
    _apply_color_scale(ws, irr_range)
    row += 1

    # ── EM grid header ──────────────────────────────────────────────
    ws.cell(row=row, column=1, value="EQUITY MULTIPLE")
    _style_header_row(ws, row, len(cap_axis) + 2)
    row += 1
    ws.cell(row=row, column=1, value="NOI Multiplier \\ Exit Cap")
    for j, cap in enumerate(cap_axis, start=2):
        c = ws.cell(row=row, column=j, value=cap)
        c.number_format = PCT_FMT
        c.font = Font(bold=True)
    _style_subhead_row(ws, row, len(cap_axis) + 1)
    row += 1
    em_start_row = row
    for nm in noi_axis:
        c = ws.cell(row=row, column=1, value=nm)
        c.number_format = MULT_FMT
        c.font = Font(bold=True)
        for j, cap in enumerate(cap_axis, start=2):
            cell_data = cell_lookup.get((cap, nm))
            if cell_data is None:
                continue
            em = float(cell_data.get("equity_multiple") or 0)
            cell = ws.cell(row=row, column=j, value=em)
            cell.number_format = MULT_FMT
            if cell_data.get("breaches_dscr_floor"):
                cell.fill = dscr_red_fill
                cell.border = dscr_red_border
        row += 1
    em_end_row = row - 1
    em_range = f"{irr_first_col_letter}{em_start_row}:{irr_last_col_letter}{em_end_row}"
    _apply_color_scale(ws, em_range)

    _autosize(ws, [26] + [14] * len(cap_axis))
    _freeze_top(ws, row=2)
    return True


def _build_comparable_sales(wb: Workbook, comp_sales: dict[str, Any] | None) -> bool:
    """Comparable Sales sheet — the comp universe with derivation summary."""
    if not comp_sales:
        return False
    txns = comp_sales.get("transactions") or []
    if not txns:
        return False

    ws = wb.create_sheet("Comparable Sales")

    # ── Summary callouts ────────────────────────────────────────────
    ws["A1"] = "DERIVED EXIT CAP RATE"
    _style_header_row(ws, 1, 6)
    row = 2
    median_cap = comp_sales.get("derived_cap_rate_median")
    weighted_cap = comp_sales.get("derived_cap_rate_weighted")
    method = comp_sales.get("derived_cap_rate_method", "none")
    coverage = comp_sales.get("coverage_quality", "low")

    ws.cell(row=row, column=1, value="Method").font = Font(bold=True)
    ws.cell(row=row, column=2, value=method.title())
    ws.cell(row=row, column=3, value="Coverage").font = Font(bold=True)
    ws.cell(row=row, column=4, value=coverage.title())
    row += 1
    ws.cell(row=row, column=1, value="Median Cap").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=(median_cap or 0) / 100.0)
    c.number_format = PCT_FMT
    ws.cell(row=row, column=3, value="Weighted Cap").font = Font(bold=True)
    c = ws.cell(row=row, column=4, value=(weighted_cap or 0) / 100.0)
    c.number_format = PCT_FMT
    row += 2

    # ── Weighting notes callout ─────────────────────────────────────
    ws.cell(row=row, column=1, value="WEIGHTING NOTES")
    _style_subhead_row(ws, row, 6)
    row += 1
    for note in comp_sales.get("weighting_notes") or []:
        ws.cell(row=row, column=1, value=str(note)).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1

    # ── Transactions table ──────────────────────────────────────────
    headers = [
        "Property", "City", "Date", "Keys", "Price",
        "Per Key", "NOI", "Cap %", "Chain", "Excluded?",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1
    excluded_fill = PatternFill("solid", fgColor="FEE2E2")
    for txn in txns:
        ws.cell(row=row, column=1, value=txn.get("property_name") or "")
        ws.cell(row=row, column=2, value=txn.get("city") or "")
        date_val = txn.get("sale_date") or ""
        ws.cell(row=row, column=3, value=str(date_val))
        c = ws.cell(row=row, column=4, value=int(txn.get("keys") or 0))
        c.number_format = INT_FMT
        c = ws.cell(row=row, column=5, value=float(txn.get("sale_price_usd") or 0))
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=6, value=float(txn.get("sale_price_per_key_usd") or 0))
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=7, value=float(txn.get("noi_usd") or 0))
        c.number_format = USD_FMT
        cap_v = txn.get("cap_rate_pct")
        c = ws.cell(
            row=row, column=8,
            value=(float(cap_v) / 100.0) if cap_v is not None else None,
        )
        c.number_format = PCT_FMT
        ws.cell(row=row, column=9, value=txn.get("chain_scale") or "—")
        excluded = bool(txn.get("excluded"))
        ws.cell(row=row, column=10, value="Yes" if excluded else "No")
        if excluded:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = excluded_fill
        row += 1

    _autosize(ws, [32, 16, 14, 8, 16, 16, 16, 10, 16, 12])
    _freeze_top(ws, row=2)
    return True


def _build_historical_baseline(wb: Workbook, baseline: dict[str, Any] | None) -> bool:
    """Historical Baseline sheet — 5-year horizontal P&L table + YoY walk."""
    if not baseline:
        return False
    years = baseline.get("years") or []
    if not years:
        return False

    ws = wb.create_sheet("Historical Baseline")

    # ── Coverage banner ─────────────────────────────────────────────
    coverage = float(baseline.get("coverage_pct") or 0)
    look_back = int(baseline.get("look_back_years") or 5)
    ws["A1"] = f"COVERAGE: {coverage:.0%} ({len(years)} of {look_back} years)"
    _style_header_row(ws, 1, 8)
    row = 2

    gaps = baseline.get("gaps") or []
    if gaps:
        ws.cell(row=row, column=1, value=f"Missing years: {', '.join(str(g) for g in gaps)}")
        ws.cell(row=row, column=1).font = Font(italic=True, color="92400E")
        row += 1

    # ── Horizontal P&L table ────────────────────────────────────────
    row += 1
    fy_list = sorted({int(y["fiscal_year"]) for y in years if y.get("fiscal_year")})
    year_lookup = {int(y["fiscal_year"]): y for y in years if y.get("fiscal_year")}

    headers = ["Line Item"] + [f"FY{fy}" for fy in fy_list]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    p_and_l_lines = (
        ("rooms_revenue", "Rooms Revenue", USD_FMT),
        ("fnb_revenue", "F&B Revenue", USD_FMT),
        ("other_revenue", "Other Revenue", USD_FMT),
        ("total_revenue", "Total Revenue", USD_FMT),
        ("gop", "GOP", USD_FMT),
        ("noi", "NOI", USD_FMT),
        ("occupancy", "Occupancy", PCT_FMT),
        ("adr", "ADR", USD_FMT),
        ("revpar", "RevPAR", USD_FMT),
    )
    for key, label, fmt in p_and_l_lines:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        for i, fy in enumerate(fy_list, start=2):
            v = year_lookup.get(fy, {}).get(key)
            if v is None:
                continue
            c = ws.cell(row=row, column=i, value=float(v))
            c.number_format = fmt
        row += 1

    # ── YoY walk — top swings ───────────────────────────────────────
    row += 1
    walk = baseline.get("walk") or []
    if walk:
        ws.cell(row=row, column=1, value="TOP YoY SWINGS")
        _style_subhead_row(ws, row, 4)
        row += 1
        ws.cell(row=row, column=1, value="Line")
        ws.cell(row=row, column=2, value="Year")
        ws.cell(row=row, column=3, value="Value")
        ws.cell(row=row, column=4, value="YoY %")
        _style_header_row(ws, row, 4)
        row += 1
        walk_start = row
        for d in walk[:5]:
            ws.cell(row=row, column=1, value=str(d.get("line", "")).replace("_", " ").title())
            ws.cell(row=row, column=2, value=int(d.get("year") or 0))
            c = ws.cell(row=row, column=3, value=float(d.get("value") or 0))
            c.number_format = USD_FMT
            pct = float(d.get("yoy_pct") or 0)
            c = ws.cell(row=row, column=4, value=pct)
            c.number_format = PCT_FMT
            row += 1
        walk_end = row - 1
        if walk_end >= walk_start:
            _apply_color_scale(
                ws,
                f"D{walk_start}:D{walk_end}",
                lo_color="FECACA",
                mid_color="FEF3C7",
                hi_color="BBF7D0",
            )

    _autosize(ws, [26] + [16] * max(1, len(fy_list)))
    _freeze_top(ws, row=2)
    return True


def _build_str_forecast(wb: Workbook, str_forecast: dict[str, Any] | None) -> bool:
    """STR Forward Forecast sheet — 48 rows (24 historical + 24 forecast).

    Cols: Month, Historical RevPAR, Downside RevPAR, Base RevPAR,
    Upside RevPAR, Comp Set RevPAR, Subject Index.
    """
    if not str_forecast:
        return False

    historical = str_forecast.get("historical_months") or []
    forecast = str_forecast.get("forecast_months") or {}
    if not historical and not forecast:
        return False

    ws = wb.create_sheet("STR Forecast")
    headers = [
        "Month", "Historical RevPAR", "Downside RevPAR", "Base RevPAR",
        "Upside RevPAR", "Comp Set RevPAR", "Subject Index",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    row = 2
    # Historical rows
    for m in historical:
        ws.cell(row=row, column=1, value=str(m.get("period") or ""))
        c = ws.cell(row=row, column=2, value=float(m.get("revpar") or 0))
        c.number_format = USD_FMT
        # Forecast columns left blank for historical rows.
        c = ws.cell(row=row, column=6, value=float(m.get("comp_set_revpar") or 0))
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=7, value=float(m.get("revpar_index") or 0))
        c.number_format = MULT_FMT
        row += 1

    # Forecast rows — align downside / base / upside side-by-side by period.
    downside = forecast.get("downside") or []
    base = forecast.get("base") or []
    upside = forecast.get("upside") or []
    n = max(len(downside), len(base), len(upside))
    for i in range(n):
        d = downside[i] if i < len(downside) else {}
        b = base[i] if i < len(base) else {}
        u = upside[i] if i < len(upside) else {}
        period = b.get("period") or d.get("period") or u.get("period") or ""
        ws.cell(row=row, column=1, value=str(period))
        # Historical RevPAR column intentionally blank for forecast rows.
        if d:
            c = ws.cell(row=row, column=3, value=float(d.get("revpar") or 0))
            c.number_format = USD_FMT
        if b:
            c = ws.cell(row=row, column=4, value=float(b.get("revpar") or 0))
            c.number_format = USD_FMT
        if u:
            c = ws.cell(row=row, column=5, value=float(u.get("revpar") or 0))
            c.number_format = USD_FMT
        # Comp set + index from base scenario (canonical).
        if b:
            c = ws.cell(row=row, column=6, value=float(b.get("comp_set_revpar") or 0))
            c.number_format = USD_FMT
            c = ws.cell(row=row, column=7, value=float(b.get("revpar_index") or 0))
            c.number_format = MULT_FMT
        row += 1

    _autosize(ws, [12, 18, 18, 18, 18, 18, 14])
    _freeze_top(ws, row=2)
    return True


def _build_named_scenarios(
    wb: Workbook, scenarios: list[dict[str, Any]] | None
) -> bool:
    """Named Scenarios Compare sheet.

    Rows = KPIs (IRR, EM, NOI Y1, NOI Stab, Exit Cap, DSCR).
    Cols = scenarios. % delta from base in adjacent cells.
    """
    if not scenarios or len(scenarios) < 2:
        return False

    ws = wb.create_sheet("Named Scenarios")
    base = next((s for s in scenarios if s.get("is_base")), scenarios[0])
    others = [s for s in scenarios if s is not base]

    # Header row: Scenario columns
    ws.cell(row=1, column=1, value="KPI")
    col = 2
    for s in [base] + others:
        ws.cell(row=1, column=col, value=str(s.get("name") or "—"))
        col += 1
        ws.cell(row=1, column=col, value="Δ vs Base")
        col += 1
    _style_header_row(ws, 1, col - 1)

    # Description row (optional)
    ws.cell(row=2, column=1, value="Description").font = Font(italic=True)
    col = 2
    for s in [base] + others:
        cell = ws.cell(row=2, column=col, value=s.get("description") or "")
        cell.alignment = Alignment(wrap_text=True)
        cell.font = Font(italic=True, color="6B7280")
        col += 2

    kpis = (
        ("levered_irr", "Levered IRR", PCT_FMT),
        ("equity_multiple", "Equity Multiple", MULT_FMT),
        ("year1_noi_usd", "Year 1 NOI", USD_FMT),
        ("stabilized_noi_usd", "Stabilized NOI", USD_FMT),
        ("exit_cap_pct", "Exit Cap Rate", PCT_FMT),
        ("year1_dscr", "Year 1 DSCR", MULT_FMT),
    )

    base_kpis = base.get("kpis") or {}
    row = 3
    for key, label, fmt in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        base_val = float(base_kpis.get(key) or 0)
        col = 2
        for s in [base] + others:
            k = s.get("kpis") or {}
            v = float(k.get(key) or 0)
            c = ws.cell(row=row, column=col, value=v)
            c.number_format = fmt
            col += 1
            # Delta column (blank for the base scenario itself)
            if s is base:
                ws.cell(row=row, column=col, value=None)
            else:
                if base_val == 0:
                    delta = 0.0
                else:
                    delta = (v - base_val) / abs(base_val)
                c = ws.cell(row=row, column=col, value=delta)
                c.number_format = PCT_FMT
            col += 1
        row += 1

    widths = [22] + [14, 12] * (1 + len(others))
    _autosize(ws, widths)
    _freeze_top(ws, row=2)
    return True


def _build_loi_appendix(wb: Workbook, loi: dict[str, Any] | None) -> bool:
    """LOI Appendix sheet — plain-text markdown body in a wide column.

    Each markdown line lands on its own row in column A. Headings get
    bold formatting; ``---`` rules are skipped.
    """
    if not loi:
        return False
    md = loi.get("rendered_markdown") or ""
    if not md.strip():
        return False

    ws = wb.create_sheet("LOI Appendix")
    ws["A1"] = "DRAFT LETTER OF INTENT"
    _style_header_row(ws, 1, 1)

    row = 2
    for raw in md.splitlines():
        line = raw.rstrip()
        if line.strip() == "---":
            row += 1
            continue
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("# "):
            cell.value = line[2:].strip()
            cell.font = Font(bold=True, size=14)
        elif line.startswith("## "):
            cell.value = line[3:].strip()
            cell.font = Font(bold=True, size=12)
        elif line.startswith("### "):
            cell.value = line[4:].strip()
            cell.font = Font(bold=True, size=11)
        row += 1

    ws.column_dimensions["A"].width = 100
    return True


# ─────────────────────────── public API ───────────────────────────


def build_excel(deal_id: UUID | str, model: dict[str, Any], output_path: Path) -> Path:
    """Build the institutional underwriting workbook.

    Always renders the 9 legacy sheets (Cover/Assumptions/Sources &
    Uses/Operating Proforma/Debt Schedule/Returns/Partnership/Variance/
    Market Comps). The legacy Sensitivity sheet is REPLACED by the W4.2
    Pricing Sensitivity sheet when ``sensitivity_grid`` is set (when
    absent, no sensitivity sheet is emitted — barebones deals get the
    9 legacy sheets cleanly).

    Wave 2/3 sheets render conditionally on presence of their source
    data, so a barebones deal still produces a valid workbook.
    """
    wave = _aggregate_wave2_3_for_excel(model)

    wb = Workbook()
    # Reserve the active sheet for Cover — we'll mutate it at the end
    # once we know which Wave 2/3 sheets actually rendered.
    cover_ws = wb.active
    cover_ws.title = "Cover"

    sections_included: list[str] = ["Cover"]

    _build_assumptions(wb, model)
    sections_included.append("Assumptions")
    _build_sources_uses(wb, model)
    sections_included.append("Sources & Uses")
    _build_proforma(wb, model)
    sections_included.append("Operating Proforma")
    _build_debt_schedule(wb, model)
    sections_included.append("Debt Schedule")
    _build_returns(wb, model)
    sections_included.append("Returns")
    _build_partnership(wb, model)
    sections_included.append("Partnership")
    _build_variance(wb, model)
    sections_included.append("Variance")
    _build_market_comps(wb, model)
    sections_included.append("Market Comps")

    # Wave 2/3 sheets — only when data present.
    if _build_revenue_mix(wb, wave["segments_by_year"]):
        sections_included.append("Revenue Mix")
    if _build_renovation_plan(wb, wave["pip"]):
        sections_included.append("Renovation Plan")
    if _build_capital_plan(wb, wave["capex_schedule"]):
        sections_included.append("Capital Plan")
    if _build_op_ratio_provenance(wb, wave["op_ratio_provenance"]):
        sections_included.append("Op-Ratio Provenance")
    # Glue the max-price summary into the sensitivity grid sheet too.
    if wave["sensitivity_grid"] and wave["max_price"]:
        wave["sensitivity_grid"]["max_price"] = wave["max_price"]
    if _build_sensitivity_grid(wb, wave["sensitivity_grid"]):
        sections_included.append("Pricing Sensitivity")
    if _build_comparable_sales(wb, wave["comp_sales"]):
        sections_included.append("Comparable Sales")
    if _build_historical_baseline(wb, wave["historical_baseline"]):
        sections_included.append("Historical Baseline")
    if _build_str_forecast(wb, wave["str_forecast"]):
        sections_included.append("STR Forecast")
    if _build_named_scenarios(wb, wave["named_scenarios"]):
        sections_included.append("Named Scenarios")
    if _build_loi_appendix(wb, wave["loi"]):
        sections_included.append("LOI Appendix")

    # Now render the Cover with the actual section list. The active
    # sheet is still our reserved Cover blank, so _build_cover mutates
    # it in place.
    _build_cover(wb, model, sections_included)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


__all__ = [
    "_aggregate_wave2_3_for_excel",
    "_apply_color_scale",
    "build_excel",
]
