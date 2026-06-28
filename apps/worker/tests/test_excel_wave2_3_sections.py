"""Wave 4 W4.2 — Excel acquisition model refresh tests.

Mirrors ``test_memo_pdf_wave2_sections.py`` for the .xlsx artifact. The
W4.2 refresh added 10 Wave 2/3 conditional sheets on top of the 9
always-on legacy sheets. The most important property is **backward
compat**: a deal with no Wave 2/3 data on the model must still produce
a valid workbook (legacy 9-sheet shape) without leaking any of the new
sheets through.

Test catalog (14 tests, mirrors the W4.2 spec):

    1.  Barebones deal → valid xlsx + no Wave 2/3 sheets
    2.  Revenue Mix present when segments populated
    3.  Revenue Mix absent when segments empty
    4.  Renovation Plan only when pip displacement set (non-"none")
    5.  Capital Plan three-bucket includes phasing
    6.  Op-Ratio Provenance lists winning source per ratio
    7.  Pricing Sensitivity has 5x5 grid + color scale
    8.  Pricing Sensitivity marks DSCR-breach cells
    9.  Comp Sales present when transactions exist
    10. Historical Baseline present when coverage high
    11. STR Forecast has 48 data rows
    12. Named Scenarios Compare present when scenarios exist
    13. LOI Appendix renders plain text
    14. Cover sheet section list omits empty sections
"""

from __future__ import annotations

import copy
import os
import shutil
import tempfile
from pathlib import Path

import pytest

os.environ.setdefault("DATABASE_URL", "sqlite+aiosqlite:///./fondok.db")


@pytest.fixture
def tmp_out() -> Path:
    d = Path(tempfile.mkdtemp(prefix="fondok-excel-wave23-"))
    yield d
    shutil.rmtree(d, ignore_errors=True)


# ─────────────────────── helpers ───────────────────────


def _barebones_model() -> dict:
    """Pre-Wave-2 minimal model — no segments / pip / capex etc.

    The legacy 9-sheet workbook (Cover + Assumptions + Sources & Uses
    + Operating Proforma + Debt Schedule + Returns + Partnership +
    Variance + Market Comps) should still render off this dict.
    """
    return {
        "deal_id": "barebones-test",
        "deal_name": "Barebones Test Hotel",
        "location": "Nowhere, USA",
        "brand": "Independent",
        "keys": 100,
        "investment_engine": {
            "purchase_price_usd": 10_000_000,
            "price_per_key_usd": 100_000,
            "total_capital_usd": 12_000_000,
            "entry_cap_rate_year1_uw": 0.08,
        },
        "p_and_l_engine_proforma": {
            "lines": [
                {"label": "Room Revenue", "y1": 8000, "y2": 8200, "y3": 8400,
                 "y4": 8600, "y5": 8800},
                {"label": "Total Revenue", "y1": 10000, "y2": 10300, "y3": 10600,
                 "y4": 10900, "y5": 11200, "bold": True},
                {"label": "Net Operating Income", "y1": 2000, "y2": 2100,
                 "y3": 2200, "y4": 2300, "y5": 2400, "bold": True},
            ],
        },
        "debt_engine": {
            "loan_amount_usd": 7_200_000,
            "interest_rate_pct": 0.06,
            "amortization_years": 30,
            "term_years": 5,
            "year1_dscr": 1.45,
            "year1_debt_yield": 0.16,
        },
        "returns_engine": {
            "hold_years": 5,
            "levered_irr": 0.12,
            "equity_multiple": 1.7,
            "year1_cash_on_cash": 0.03,
            "exit_cap_rate_pct": 0.075,
        },
        "cash_flow_engine": {},
        "partnership_engine": {},
    }


def _build_to(tmp_out: Path, model: dict) -> Path:
    from app.export.excel import build_excel
    out = tmp_out / "model.xlsx"
    build_excel("test-deal", model, out)
    return out


# ──────────────────────── tests ────────────────────────


def test_barebones_deal_still_produces_valid_xlsx(tmp_out: Path) -> None:
    """Backward compat — a deal with no Wave 2/3 data on the model still
    produces a valid workbook. NONE of the new sheets leak through."""
    from openpyxl import load_workbook

    out = _build_to(tmp_out, _barebones_model())
    assert out.exists()
    assert out.stat().st_size > 6_000, "barebones xlsx is suspiciously small"

    wb = load_workbook(out, read_only=True)
    # All 9 legacy sheets are present.
    legacy = {
        "Cover", "Assumptions", "Sources & Uses", "Operating Proforma",
        "Debt Schedule", "Returns", "Partnership", "Variance", "Market Comps",
    }
    assert legacy.issubset(set(wb.sheetnames)), (
        f"missing legacy sheets: {legacy - set(wb.sheetnames)}"
    )
    # NONE of the Wave 2/3 sheets leaked through.
    wave = {
        "Revenue Mix", "Renovation Plan", "Capital Plan",
        "Op-Ratio Provenance", "Pricing Sensitivity", "Comparable Sales",
        "Historical Baseline", "STR Forecast", "Named Scenarios",
        "LOI Appendix",
    }
    leaked = wave & set(wb.sheetnames)
    assert not leaked, f"Wave 2/3 sheets leaked into barebones workbook: {leaked}"
    # Exactly 9 sheets.
    assert len(wb.sheetnames) == 9, f"got {len(wb.sheetnames)}: {wb.sheetnames}"
    wb.close()


def test_revenue_mix_sheet_present_when_segments_populated(tmp_out: Path) -> None:
    """Revenue Mix sheet renders rows for each segment + totals.

    Kimpton fixture carries 5 demand segments across 5 hold years
    (W3.4 → W4.2 enrichment), so the sheet should show 25+ segment rows
    plus year-total and grand-total rows.
    """
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    assert "Revenue Mix" in wb.sheetnames

    ws = wb["Revenue Mix"]
    # Collect column A values
    col_a = [row[0].value for row in ws.iter_rows(values_only=False)]
    # Header row + Year subheads + segment rows + total rows
    seg_names = {"transient_bar", "transient_ota", "corporate", "group", "contract"}
    assert seg_names.issubset({v for v in col_a if v in seg_names})
    # Grand total appears at the bottom.
    assert any(v == "GRAND TOTAL" for v in col_a)
    wb.close()


def test_revenue_mix_sheet_absent_when_segments_empty(tmp_out: Path) -> None:
    """Empty / missing segments_by_year → no Revenue Mix sheet."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    model = copy.deepcopy(kimpton_model())
    model["segments_by_year"] = []
    if "revenue_engine" in model:
        model["revenue_engine"].pop("segments_by_year", None)
        model["revenue_engine"].pop("segment_breakdown", None)

    out = _build_to(tmp_out, model)
    wb = load_workbook(out, read_only=True)
    assert "Revenue Mix" not in wb.sheetnames
    wb.close()


def test_renovation_plan_sheet_present_only_when_pip_displacement_set(
    tmp_out: Path,
) -> None:
    """Renovation Plan sheet renders only when pip displacement carries
    a non-"none" closure_strategy."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    # Case 1: fixture has rolling closure → sheet present
    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    assert "Renovation Plan" in wb.sheetnames
    ws = wb["Renovation Plan"]
    # Strategy row should carry the "Rolling" label.
    all_vals = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("Rolling" == v for v in all_vals)
    wb.close()

    # Case 2: closure_strategy = "none" → sheet absent
    model_none = copy.deepcopy(kimpton_model())
    model_none["pip_displacement"] = {
        "closure_strategy": "none",
        "pct_rooms_offline_by_month": [],
    }
    out2 = tmp_out / "no_pip.xlsx"
    from app.export.excel import build_excel
    build_excel("no-pip-deal", model_none, out2)
    wb2 = load_workbook(out2, read_only=True)
    assert "Renovation Plan" not in wb2.sheetnames
    wb2.close()

    # Case 3: pip_displacement missing → sheet absent
    model_missing = copy.deepcopy(kimpton_model())
    model_missing.pop("pip_displacement", None)
    out3 = tmp_out / "missing_pip.xlsx"
    build_excel("missing-pip-deal", model_missing, out3)
    wb3 = load_workbook(out3, read_only=True)
    assert "Renovation Plan" not in wb3.sheetnames
    wb3.close()


def test_capital_plan_three_bucket_sheet_includes_phasing(tmp_out: Path) -> None:
    """Capital Plan sheet renders 5 hold years × PIP/Non-PIP/ROI cols
    plus a Total row."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    assert "Capital Plan" in wb.sheetnames
    ws = wb["Capital Plan"]
    headers = [c.value for c in ws[1]]
    assert "PIP" in headers
    assert "Non-PIP FF&E" in headers
    assert "ROI Investment" in headers
    assert "ROI NOI Lift" in headers
    assert "Total Capex" in headers
    # Pull column A values
    col_a = [c.value for row in ws.iter_rows() for c in row[:1]]
    # Years 1..5 + a Total row.
    for y in range(1, 6):
        assert f"Year {y}" in col_a
    assert "Total" in col_a
    wb.close()


def test_op_ratio_provenance_sheet_lists_winning_source_per_ratio(
    tmp_out: Path,
) -> None:
    """Op-Ratio Provenance sheet has 1 row per ratio with the winning
    source labelled (T-12, Portfolio, CBRE, HOST, or Override)."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    assert "Op-Ratio Provenance" in wb.sheetnames
    ws = wb["Op-Ratio Provenance"]
    # Collect source-column values (col 3).
    sources = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
    # The fixture has every source represented at least once.
    seen = {s for s in sources if s}
    expected_labels = {"T-12", "Portfolio", "CBRE", "HOST", "Override"}
    assert expected_labels.issubset(seen), (
        f"missing source labels {expected_labels - seen}"
    )
    wb.close()


def test_sensitivity_sheet_has_5x5_grid_and_color_scale(tmp_out: Path) -> None:
    """Pricing Sensitivity sheet carries a 5×5 IRR grid + 5×5 EM grid
    with a ColorScaleRule attached (real conditional format, not just
    baked-in fill)."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out)  # need write-mode for conditional_formatting access
    assert "Pricing Sensitivity" in wb.sheetnames
    ws = wb["Pricing Sensitivity"]

    # The fixture has 5 cap-rate cols × 5 NOI-multiplier rows.
    # The sheet sits IRR grid then EM grid below — we look for 5 axis
    # cells (NOI multipliers, column A) per grid, so 10 across both.
    col_a_vals = [
        ws.cell(row=r, column=1).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value is not None
    ]
    # Count distinct NOI multiplier axis values that show up at least twice
    # (once in IRR grid, once in EM grid).
    mult_vals = [v for v in col_a_vals if isinstance(v, (int, float))]
    counts: dict = {}
    for v in mult_vals:
        counts[v] = counts.get(v, 0) + 1
    paired = [v for v, c in counts.items() if c >= 2]
    assert len(paired) >= 5, f"expected 5 paired NOI axis values, got {paired}"

    # ColorScale rule check — iterate the sheet's conditional_formatting.
    # ``ColorScaleRule(...)`` is a factory that returns an
    # ``openpyxl.formatting.rule.Rule`` with ``type='colorScale'`` and
    # ``colorScale`` populated. We assert on those attrs rather than on
    # the class (ColorScaleRule is a function, not a class).
    rules: list = []
    for _cf_range, cf_rules in ws.conditional_formatting._cf_rules.items():
        rules.extend(cf_rules)
    color_scale_rules = [
        r for r in rules
        if getattr(r, "type", None) == "colorScale"
        and getattr(r, "colorScale", None) is not None
    ]
    assert color_scale_rules, (
        "Pricing Sensitivity sheet should have at least one colorScale rule"
    )
    # Expect at least 2 — one for IRR grid + one for EM grid.
    assert len(color_scale_rules) >= 2, (
        f"expected ≥2 colorScale rules (IRR + EM grids), got {len(color_scale_rules)}"
    )
    wb.close()


def test_sensitivity_sheet_marks_dscr_breach_cells(tmp_out: Path) -> None:
    """DSCR-breach cells get a red fill + ``!`` comment override on top
    of the color-scale rule."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out)
    ws = wb["Pricing Sensitivity"]

    # Walk the IRR grid looking for red-fill cells with a ``!`` comment.
    breach_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment is not None and "DSCR breach" in (cell.comment.text or ""):
                breach_count += 1
    # Fixture has 5 DSCR-breach cells (NOI 0.85x row across 5 cap rates).
    # In the IRR grid AND the EM grid the borders fire, but only the
    # IRR grid carries the ``!`` comment. So we expect ≥ 5 here.
    assert breach_count >= 5, f"expected ≥5 DSCR-breach comments, got {breach_count}"
    wb.close()


def test_comp_sales_sheet_present_when_transactions_exist(tmp_out: Path) -> None:
    """Comparable Sales sheet renders the comp universe + median +
    weighted cap callouts when comp_sales has ≥1 transaction."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    assert "Comparable Sales" in wb.sheetnames
    ws = wb["Comparable Sales"]
    all_vals = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    # Headline callouts.
    assert any("Median Cap" == v for v in all_vals)
    assert any("Weighted Cap" == v for v in all_vals)
    # At least one comp property name appears.
    assert any(isinstance(v, str) and "Setai" in v for v in all_vals)
    wb.close()

    # When comp_sales is empty → sheet absent
    model_empty = copy.deepcopy(kimpton_model())
    model_empty.pop("comp_sales", None)
    out2 = tmp_out / "no_comps.xlsx"
    from app.export.excel import build_excel
    build_excel("no-comps-deal", model_empty, out2)
    wb2 = load_workbook(out2, read_only=True)
    assert "Comparable Sales" not in wb2.sheetnames
    wb2.close()


def test_historical_baseline_sheet_present_when_coverage_high(tmp_out: Path) -> None:
    """Historical Baseline renders the per-year P&L table + YoY walk
    when coverage_pct > 0."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    assert "Historical Baseline" in wb.sheetnames
    ws = wb["Historical Baseline"]
    headers = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    # Coverage banner.
    assert any(isinstance(v, str) and v.startswith("COVERAGE:") for v in headers)
    # Per-line labels — check NOI + Rooms Revenue + RevPAR rows.
    expected_labels = {"NOI", "Rooms Revenue", "RevPAR"}
    line_labels = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
    assert expected_labels.issubset(line_labels)
    # YoY walk section header.
    assert any(v == "TOP YoY SWINGS" for v in headers)
    wb.close()


def test_str_forecast_sheet_has_48_data_rows(tmp_out: Path) -> None:
    """STR Forecast sheet has 48 data rows (24 historical + 24 forecast)."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    assert "STR Forecast" in wb.sheetnames
    ws = wb["STR Forecast"]
    # Count rows in column A from row 2 (header is row 1) that have a value.
    data_rows = [
        ws.cell(row=r, column=1).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    ]
    assert len(data_rows) == 48, f"expected 48 data rows, got {len(data_rows)}"
    wb.close()


def test_named_scenarios_compare_sheet_present_when_scenarios_exist(
    tmp_out: Path,
) -> None:
    """Named Scenarios sheet renders one column per scenario + delta
    cols, with KPI rows (IRR, EM, Y1 NOI, Stab NOI, Exit Cap, DSCR)."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    assert "Named Scenarios" in wb.sheetnames
    ws = wb["Named Scenarios"]
    # Column A holds the KPI labels — the 6 we care about are present.
    col_a = {ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)}
    expected = {
        "Levered IRR", "Equity Multiple", "Year 1 NOI",
        "Stabilized NOI", "Exit Cap Rate", "Year 1 DSCR",
    }
    assert expected.issubset(col_a), f"missing KPI rows: {expected - col_a}"
    # Header row has scenario names.
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Base Case" in header
    assert "PIP Skinny" in header
    wb.close()

    # Single scenario (or none) → sheet absent
    model_one = copy.deepcopy(kimpton_model())
    model_one["named_scenarios"] = [model_one["named_scenarios"][0]]
    out2 = tmp_out / "one_scenario.xlsx"
    from app.export.excel import build_excel
    build_excel("one-scenario", model_one, out2)
    wb2 = load_workbook(out2, read_only=True)
    assert "Named Scenarios" not in wb2.sheetnames
    wb2.close()


def test_loi_appendix_sheet_renders_plain_text(tmp_out: Path) -> None:
    """LOI Appendix sheet renders the LOI markdown body, one line per row,
    with headings styled bold."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    assert "LOI Appendix" in wb.sheetnames
    ws = wb["LOI Appendix"]
    # Concatenate column A.
    full_text = "\n".join(
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    )
    assert "LETTER OF INTENT" in full_text
    assert "Earnest Money" in full_text or "Earnest" in full_text
    # Headline price callout from the fixture.
    assert "$42,800,000" in full_text
    wb.close()


def test_cover_sheet_section_list_omits_empty_sections(tmp_out: Path) -> None:
    """Cover sheet's 'Sections Included' list reflects ONLY the sheets
    that actually rendered for this deal."""
    from openpyxl import load_workbook
    from app.export.fixtures import kimpton_model

    # Kimpton fixture has every Wave 2/3 section → list shows them all.
    out = _build_to(tmp_out, kimpton_model())
    wb = load_workbook(out, read_only=True)
    ws = wb["Cover"]
    all_vals = [
        str(ws.cell(row=r, column=1).value or "")
        for r in range(1, ws.max_row + 1)
    ]
    full = "\n".join(all_vals)
    # Should list each rendered Wave 2/3 sheet.
    assert "SECTIONS INCLUDED" in full
    for name in (
        "Revenue Mix", "Renovation Plan", "Capital Plan",
        "Op-Ratio Provenance", "Pricing Sensitivity", "Comparable Sales",
        "Historical Baseline", "STR Forecast", "Named Scenarios",
        "LOI Appendix",
    ):
        assert name in full, f"Cover section list missing {name}"
    wb.close()

    # Barebones model → Cover lists only the 9 legacy sections.
    out2 = tmp_out / "bare.xlsx"
    from app.export.excel import build_excel
    build_excel("bare", _barebones_model(), out2)
    wb2 = load_workbook(out2, read_only=True)
    ws2 = wb2["Cover"]
    full2 = "\n".join(
        str(ws2.cell(row=r, column=1).value or "")
        for r in range(1, ws2.max_row + 1)
    )
    # None of the Wave 2/3 names should appear in the section list.
    # We need to be careful not to false-match on the "QUICK READ" /
    # "KEY METRICS" sections. We look for the SECTIONS INCLUDED header
    # and the lines following it.
    idx = full2.find("SECTIONS INCLUDED")
    assert idx >= 0, "Cover sheet missing 'SECTIONS INCLUDED' header"
    tail = full2[idx:]
    for name in (
        "Revenue Mix", "Renovation Plan", "Capital Plan",
        "Op-Ratio Provenance", "Pricing Sensitivity", "Comparable Sales",
        "Historical Baseline", "STR Forecast", "Named Scenarios",
        "LOI Appendix",
    ):
        assert name not in tail, (
            f"barebones Cover should not list {name!r}, but it appears in: {tail!r}"
        )
    wb2.close()
