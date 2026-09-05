"""FON-54 follow-up — the LIVE export payload carries the Wave 2/3 sheets.

``load_live_payload`` previously assembled only the 8-engine canonical
snapshot (+ Revenue Mix) for real UUID deals, so a real deal's Excel model
shipped without Variance / Market Comps rows and without the Named Scenarios,
Capital Plan, Pricing Sensitivity, Historical Baseline and LOI sheets the
Kimpton fixture carries. These tests lock:

    1. a seeded real-UUID deal with a canonical run + saved scenarios gets
       ``named_scenarios`` / ``scenario_outputs`` (Base pinned to the
       canonical run, not the downside scenario), ``capex_schedule`` from the
       live capital + expense outputs, and the pricing trio
       (``sensitivity_grid`` / ``max_price`` / ``loi_draft``) — all with real
       values and NO fixture strings;
    2. sheets whose live source is empty are OMITTED (never fixture rows);
    3. transaction comps extracted on the deal populate ``market_comps``;
    4. extracted P&L years populate ``historical_baseline`` + its YoY walk;
    5. the variance-flag mapper preserves the UI's severity / values;
    6. the slug/demo Kimpton path is untouched.
"""

from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path
from uuid import uuid4

import pytest
from sqlalchemy import text

# Force a per-test SQLite DB BEFORE app modules import so the cached
# Settings/engine pick up the right DSN (same pattern as test_fon73_*).
_TMP_DB = Path(tempfile.gettempdir()) / "fondok-tests-live-payload-w23.db"
if _TMP_DB.exists():
    _TMP_DB.unlink()
os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///{_TMP_DB}"

# Strings that only exist in the Kimpton fixture — none may leak into a real
# deal's export payload.
_FIXTURE_STRINGS = (
    "Kimpton",
    "Miami Beach",
    "The Setai",
    "Nautilus by Arlo",
    "T12_FinancialStatement.xlsx",
    "Offering_Memorandum_Final.pdf",
    "Monthly_PL_2022.xlsx",
    "PIP Skinny",
    "Aggressive Exit",
    "660 Washington Ave",
)

_DOWNSIDE = {"hold_years": 7, "starting_occupancy": 0.55}


@pytest.fixture(autouse=True)
async def _reset_db() -> None:
    """Recreate + clear the schema before each test."""
    from app.database import get_session_factory
    from app.migrations import run_startup_migrations

    await run_startup_migrations()
    factory = get_session_factory()
    async with factory() as session:
        for tbl in (
            "engine_outputs",
            "scenarios",
            "extraction_results",
            "documents",
            "deals",
        ):
            try:
                await session.execute(text(f"DELETE FROM {tbl}"))
            except Exception:  # noqa: BLE001
                pass
        await session.commit()
    yield


async def _seed_deal(
    session,
    *,
    deal_id: str,
    tenant_id: str,
    name: str = "Harbor House Test",
    city: str = "Tampa, FL",
    keys: int = 150,
) -> None:
    await session.execute(
        text(
            "INSERT INTO deals (id, tenant_id, name, city, keys, field_overrides) "
            "VALUES (:id, :tenant, :name, :city, :keys, :fo)"
        ),
        {
            "id": deal_id,
            "tenant": tenant_id,
            "name": name,
            "city": city,
            "keys": keys,
            "fo": "{}",
        },
    )


async def _insert_scenario(
    session,
    *,
    deal_id: str,
    tenant_id: str,
    name: str,
    is_base: bool,
    run_id: str | None,
    description: str | None = None,
) -> str:
    sid = str(uuid4())
    await session.execute(
        text(
            "INSERT INTO scenarios (id, deal_id, tenant_id, name, description, "
            "is_base, overrides, last_run_id) VALUES "
            "(:id, :deal, :tenant, :name, :desc, :is_base, :ov, :run)"
        ),
        {
            "id": sid,
            "deal": deal_id,
            "tenant": tenant_id,
            "name": name,
            "desc": description,
            "is_base": is_base,
            "ov": "[]",
            "run": run_id,
        },
    )
    return sid


async def _seed_doc_with_fields(
    session,
    *,
    deal_id: str,
    tenant_id: str,
    filename: str,
    doc_type: str,
    fields: list[dict[str, object]],
    fiscal_year: int | None = None,
    status: str = "Extracted",
) -> str:
    doc_id = str(uuid4())
    await session.execute(
        text(
            """
            INSERT INTO documents (
                id, deal_id, tenant_id, filename, doc_type, status,
                fiscal_year, usali_deviations
            ) VALUES (
                :id, :deal, :tenant, :fname, :dtype, :stat, :year, :devs
            )
            """
        ),
        {
            "id": doc_id,
            "deal": deal_id,
            "tenant": tenant_id,
            "fname": filename,
            "dtype": doc_type,
            "stat": status,
            "year": fiscal_year,
            "devs": json.dumps(
                {"inconclusive": False, "applicable_count": 0,
                 "passed_count": 0, "deviations": []}
            ),
        },
    )
    await session.execute(
        text(
            """
            INSERT INTO extraction_results (
                id, document_id, deal_id, tenant_id, fields
            ) VALUES (:id, :doc, :deal, :tenant, :fields)
            """
        ),
        {
            "id": str(uuid4()),
            "doc": doc_id,
            "deal": deal_id,
            "tenant": tenant_id,
            "fields": json.dumps(fields),
        },
    )
    return doc_id


def _assert_no_fixture_strings(model: dict) -> None:
    blob = json.dumps(model, default=str)
    for needle in _FIXTURE_STRINGS:
        assert needle not in blob, f"fixture string {needle!r} leaked into live payload"


# ─────────────────── 1 + 2. modeled deal: scenarios / capex / pricing ───────────────────


@pytest.mark.asyncio
async def test_live_payload_wires_scenarios_capex_and_pricing() -> None:
    from app.database import get_session_factory
    from app.export.live_payload import load_live_payload
    from app.services.engine_runner import (
        _coerce_uuid,
        get_run_status,
        run_all_engines,
    )

    deal_id = str(uuid4())
    tenant_id = str(uuid4())
    base_run = str(uuid4())
    scen_run = str(uuid4())

    factory = get_session_factory()
    async with factory() as session:
        await _seed_deal(session, deal_id=deal_id, tenant_id=tenant_id)
        base = await run_all_engines(
            session, deal_id=deal_id, tenant_id=tenant_id, run_id=base_run
        )
        base_irr = base["returns"]["outputs"]["levered_irr"]

        # A LATER non-base scenario run — must not become the Base column.
        scen = await run_all_engines(
            session, deal_id=deal_id, tenant_id=tenant_id,
            run_id=scen_run, overrides=_DOWNSIDE,
        )
        scen_irr = scen["returns"]["outputs"]["levered_irr"]
        assert abs(base_irr - scen_irr) > 1e-6, "downside didn't move IRR"

        deal_uuid = str(_coerce_uuid(deal_id))
        # Base scenario row with a NULL last_run_id → export must pin it to
        # the canonical run rather than skip / stale-read it.
        await _insert_scenario(
            session, deal_id=deal_uuid, tenant_id=tenant_id,
            name="Base Case", is_base=True, run_id=None,
        )
        await _insert_scenario(
            session, deal_id=deal_uuid, tenant_id=tenant_id,
            name="Downside Stress", is_base=False, run_id=scen_run,
            description="7-yr hold, 55% starting occupancy",
        )
        # A never-run scenario is skipped (an export never triggers a run).
        await _insert_scenario(
            session, deal_id=deal_uuid, tenant_id=tenant_id,
            name="Unrun Idea", is_base=False, run_id=None,
        )
        await session.commit()

        base_rows = await get_run_status(
            session, deal_id=deal_id, run_id=base_run, tenant_id=tenant_id
        )
        deal, model, memo = await load_live_payload(session, deal_id, tenant_id)

    # ── Named Scenarios / PPTX scenario cards ──────────────────────────
    named = model["named_scenarios"]
    assert [s["name"] for s in named] == ["Base Case", "Downside Stress"]
    assert named[0]["is_base"] is True and named[1]["is_base"] is False
    assert named[0]["kpis"]["levered_irr"] == base_irr, "Base not pinned to canonical run"
    assert named[1]["kpis"]["levered_irr"] == scen_irr
    assert named[1]["description"] == "7-yr hold, 55% starting occupancy"
    for s in named:
        for k in ("levered_irr", "equity_multiple", "year1_noi_usd",
                  "stabilized_noi_usd", "exit_cap_pct", "year1_dscr"):
            assert isinstance(s["kpis"].get(k), float), f"{s['name']} missing {k}"

    outputs = model["scenario_outputs"]
    assert outputs[0]["name"] == "Base Case" and outputs[0].get("base") is True
    assert outputs[0]["irr"] == base_irr
    assert outputs[1]["irr"] == scen_irr and "base" not in outputs[1]
    assert outputs[0]["exit_value_usd"] == base["returns"]["outputs"]["gross_sale_price"]

    # ── Capital Plan from live capital + expense outputs ───────────────
    capex = model["capex_schedule"]
    hold = int(base["returns"]["outputs"]["hold_years"])
    assert [r["year"] for r in capex] == list(range(1, hold + 1))
    exp_years = {int(y["year"]): y for y in base["expense"]["outputs"]["years"]}
    for r in capex:
        assert r["non_pip_usd"] == pytest.approx(exp_years[r["year"]]["ffe_reserve"])
        assert r["roi_investment_usd"] == 0.0 and r["roi_noi_lift_usd"] == 0.0
        assert r["total_capex_usd"] == pytest.approx(r["pip_usd"] + r["non_pip_usd"])
    renovation = next(
        (u["amount"] for u in base["capital"]["outputs"]["uses"]
         if "renovation" in str(u.get("label", "")).lower() and not u.get("is_total")),
        0.0,
    )
    assert capex[0]["pip_usd"] == pytest.approx(renovation)
    assert all(r["pip_usd"] == 0.0 for r in capex[1:])
    assert sum(r["total_capex_usd"] for r in capex) > 0

    # ── Pricing trio (same math as the Pricing panel endpoints) ────────
    grid = model["sensitivity_grid"]
    assert len(grid["cells"]) == 25
    assert {round(c["exit_cap_pct"], 6) for c in grid["cells"]} and {
        c["noi_multiplier"] for c in grid["cells"]
    } == {0.85, 0.925, 1.0, 1.075, 1.15}
    base_cell = next(
        c for c in grid["cells"]
        if c["noi_multiplier"] == 1.0
        and abs(c["exit_cap_pct"] - grid["base_exit_cap_pct"]) < 1e-9
    )
    assert base_cell["levered_irr"] == pytest.approx(base_irr, abs=1e-6)
    for key in ("levered_irr", "equity_multiple", "going_in_cap_rate", "dscr_y1"):
        assert all(isinstance(c[key], float) for c in grid["cells"])

    mp = model["max_price"]
    assert mp["max_price_for_irr"] > 0 and mp["max_price_for_em"] > 0
    assert mp["binding_constraint"] in ("irr", "em", "both")
    assert mp["final_price_per_key"] == pytest.approx(
        min(mp["max_price_for_irr"], mp["max_price_for_em"]) / 150
    )

    loi = model["loi_draft"]
    assert loi["asset_name"] == "Harbor House Test"
    assert loi["rooms"] == 150
    assert loi["proposed_price"] == pytest.approx(mp["max_price_for_irr"])
    assert loi["binding_constraint"] == mp["binding_constraint"]
    assert "Harbor House Test" in loi["rendered_markdown"]
    assert "150 guest rooms" in loi["rendered_markdown"]

    # ── Sheets with no live source on this deal are OMITTED ────────────
    for absent in (
        "variance_flags", "market_comps", "comp_sales",
        "historical_baseline", "str_forecast", "pip_displacement",
        "op_ratio_provenance",
    ):
        assert absent not in model, f"{absent} must be omitted with no live data"

    _assert_no_fixture_strings(model)
    assert deal["name"] == "Harbor House Test"
    assert memo["header"]["subject_property"] == "Harbor House Test"

    # The Base column equals the canonical run's returns row byte-for-byte.
    canonical_returns = next(r for r in base_rows if r["engine"] == "returns")
    assert named[0]["kpis"]["equity_multiple"] == canonical_returns["outputs"]["equity_multiple"]


@pytest.mark.asyncio
async def test_live_payload_builds_full_excel_with_new_sheets(tmp_path: Path) -> None:
    """End-to-end: the live payload feeds ``build_excel`` and the newly wired
    sheets actually render for a real deal (and only those)."""
    from openpyxl import load_workbook

    from app.database import get_session_factory
    from app.export import build_excel
    from app.export.live_payload import load_live_payload
    from app.services.engine_runner import _coerce_uuid, run_all_engines

    deal_id = str(uuid4())
    tenant_id = str(uuid4())
    base_run = str(uuid4())
    scen_run = str(uuid4())

    factory = get_session_factory()
    async with factory() as session:
        await _seed_deal(session, deal_id=deal_id, tenant_id=tenant_id)
        await run_all_engines(session, deal_id=deal_id, tenant_id=tenant_id, run_id=base_run)
        await run_all_engines(
            session, deal_id=deal_id, tenant_id=tenant_id, run_id=scen_run, overrides=_DOWNSIDE
        )
        deal_uuid = str(_coerce_uuid(deal_id))
        await _insert_scenario(
            session, deal_id=deal_uuid, tenant_id=tenant_id,
            name="Base Case", is_base=True, run_id=base_run,
        )
        await _insert_scenario(
            session, deal_id=deal_uuid, tenant_id=tenant_id,
            name="Downside", is_base=False, run_id=scen_run,
        )
        await session.commit()
        _deal, model, _memo = await load_live_payload(session, deal_id, tenant_id)

    out = build_excel(deal_id, model, tmp_path / "live.xlsx")
    wb = load_workbook(out, read_only=True)
    names = set(wb.sheetnames)
    wb.close()

    for present in ("Named Scenarios", "Capital Plan", "Pricing Sensitivity",
                    "LOI Appendix"):
        assert present in names, f"{present} missing from live workbook: {sorted(names)}"
    # Conditional sheets whose live source is empty on a bare deal must NOT
    # render (Revenue Mix needs an STR_SEGMENTATION extraction; the rest need
    # OM comps / P&L history / STR history / a PIP spec / grounded op-ratios).
    for absent in ("Revenue Mix", "Renovation Plan", "Comparable Sales",
                   "Historical Baseline", "STR Forecast", "Op-Ratio Provenance"):
        assert absent not in names, f"{absent} must not render with no live data"
    # Always-on legacy sheets still ship (Variance / Market Comps render empty
    # tables rather than fixture rows when the deal has no flags / comps).
    assert {"Cover", "Assumptions", "Sources & Uses", "Operating Proforma",
            "Debt Schedule", "Returns", "Partnership", "Variance",
            "Market Comps"} <= names


# ─────────────────── 3. transaction comps → Market Comps rows ───────────────────


@pytest.mark.asyncio
async def test_live_payload_market_comps_from_extracted_transaction_comps() -> None:
    from app.database import get_session_factory
    from app.export.live_payload import load_live_payload

    deal_id = str(uuid4())
    tenant_id = str(uuid4())
    fields = [
        {"field_name": "transaction_comps.1.hotel_name", "value": "Bayfront Inn", "source_page": 4},
        {"field_name": "transaction_comps.1.keys", "value": 180, "source_page": 4},
        {"field_name": "transaction_comps.1.sale_date", "value": "2025-03-01", "source_page": 4},
        {"field_name": "transaction_comps.1.sale_price", "value": 54_000_000, "source_page": 4},
        {"field_name": "transaction_comps.1.cap_rate", "value": 7.1, "source_page": 4},
        {"field_name": "transaction_comps.1.buyer", "value": "Gulf Coast Lodging", "source_page": 4},
        {"field_name": "transaction_comps.2.hotel_name", "value": "Riverside Suites", "source_page": 4},
        {"field_name": "transaction_comps.2.sale_price", "value": 18_500_000, "source_page": 4},
        {"field_name": "transaction_comps.2.price_per_key", "value": 154_000, "source_page": 4},
    ]

    factory = get_session_factory()
    async with factory() as session:
        await _seed_deal(session, deal_id=deal_id, tenant_id=tenant_id)
        await _seed_doc_with_fields(
            session, deal_id=deal_id, tenant_id=tenant_id,
            filename="Bayfront_OM.pdf", doc_type="OM", fields=fields,
        )
        await session.commit()
        _deal, model, memo = await load_live_payload(session, deal_id, tenant_id)

    comps = model["market_comps"]
    assert [c["name"] for c in comps] == ["Bayfront Inn", "Riverside Suites"]
    first = comps[0]
    assert first["keys"] == 180
    assert first["date"] == "2025-03-01"
    assert first["price"] == "$54M"
    assert first["per_key"] == "$300k"  # derived 54M / 180 keys
    assert first["cap"] == "7.1%"
    assert first["buyer"] == "Gulf Coast Lodging"
    assert first["sale_price_usd"] == 54_000_000.0
    second = comps[1]
    assert second["keys"] is None, "missing keys must be blank, never 0"
    assert second["per_key"] == "$154k"
    assert second["cap"] == "" and second["buyer"] == ""

    # No modeled run on this deal → modeled sheets omitted; no fixture rows.
    for absent in ("named_scenarios", "capex_schedule", "sensitivity_grid",
                   "loi_draft", "variance_flags", "historical_baseline"):
        assert absent not in model
    _assert_no_fixture_strings(model)
    assert memo["appendix"]["documents_reviewed"] == ["Bayfront_OM.pdf"]


# ─────────────────── 4. extracted P&L years → Historical Baseline ───────────────────


def _pnl_fields(*, rooms_rev: float, fb_rev: float, noi_ish: float) -> list[dict[str, object]]:
    """Enough USALI lines for the baseline roll-up to synthesize totals."""
    return [
        {"field_name": "rooms_revenue", "value": rooms_rev, "source_page": 1},
        {"field_name": "fb_revenue", "value": fb_rev, "source_page": 1},
        {"field_name": "other_revenue", "value": 400_000.0, "source_page": 1},
        {"field_name": "rooms_dept_expense", "value": rooms_rev * 0.3, "source_page": 1},
        {"field_name": "fb_dept_expense", "value": fb_rev * 0.8, "source_page": 1},
        {"field_name": "other_dept_expense", "value": 150_000.0, "source_page": 1},
        {"field_name": "ag_expense", "value": 700_000.0, "source_page": 2},
        {"field_name": "marketing_expense", "value": 500_000.0, "source_page": 2},
        {"field_name": "utilities_expense", "value": 400_000.0, "source_page": 2},
        {"field_name": "rm_expense", "value": 350_000.0, "source_page": 2},
        {"field_name": "information_telecom", "value": 150_000.0, "source_page": 2},
        {"field_name": "property_tax", "value": 300_000.0, "source_page": 3},
        {"field_name": "insurance_expense", "value": 180_000.0, "source_page": 3},
        {"field_name": "mgmt_fee", "value": noi_ish * 0.05, "source_page": 3},
        {"field_name": "occupancy", "value": 0.71, "source_page": 1},
        {"field_name": "adr", "value": 245.0, "source_page": 1},
    ]


@pytest.mark.asyncio
async def test_live_payload_historical_baseline_from_pnl_docs() -> None:
    from app.database import get_session_factory
    from app.export.live_payload import load_live_payload

    deal_id = str(uuid4())
    tenant_id = str(uuid4())

    factory = get_session_factory()
    async with factory() as session:
        await _seed_deal(session, deal_id=deal_id, tenant_id=tenant_id)
        await _seed_doc_with_fields(
            session, deal_id=deal_id, tenant_id=tenant_id,
            filename="pnl-2023.pdf", doc_type="PNL", fiscal_year=2023,
            fields=_pnl_fields(rooms_rev=10_000_000.0, fb_rev=1_500_000.0, noi_ish=3_000_000.0),
        )
        await _seed_doc_with_fields(
            session, deal_id=deal_id, tenant_id=tenant_id,
            filename="pnl-2024.pdf", doc_type="T12", fiscal_year=2024,
            fields=_pnl_fields(rooms_rev=11_500_000.0, fb_rev=1_650_000.0, noi_ish=3_400_000.0),
        )
        await session.commit()
        _deal, model, _memo = await load_live_payload(session, deal_id, tenant_id)

    hb = model["historical_baseline"]
    assert [y["fiscal_year"] for y in hb["years"]] == [2023, 2024]
    assert hb["coverage_pct"] > 0
    assert hb["years"][0]["rooms_revenue"] == 10_000_000.0
    assert hb["years"][1]["rooms_revenue"] == 11_500_000.0
    assert hb["years"][1]["total_revenue"] == pytest.approx(11_500_000.0 + 1_650_000.0 + 400_000.0)
    # Walk rows are real YoY swings only (no first-year / no-prior rows).
    assert hb["walk"], "expected YoY walk rows for a 2-year baseline"
    assert all(d["yoy_pct"] is not None and d["year"] == 2024 for d in hb["walk"])
    rooms_walk = next(d for d in hb["walk"] if d["line"] == "rooms_revenue")
    assert rooms_walk["yoy_pct"] == pytest.approx(0.15)
    # Source docs are the real uploads, never the fixture filenames.
    assert all(
        doc_id for y in hb["years"] for doc_id in y["source_document_ids"]
    )
    _assert_no_fixture_strings(model)


# ─────────────────── 5. variance mapper (pure) ───────────────────


def test_variance_flags_mapper_preserves_ui_rows() -> None:
    from app.api.analysis import VarianceFlagOut
    from app.export.live_payload import _variance_flags_from_out

    flags = [
        VarianceFlagOut(
            field="noi", rule_id="USALI-NOI-01", severity="critical",
            actual=4_181_000.0, broker=5_200_000.0, delta=1_019_000.0,
            delta_pct=-0.196, source_page=3, note="NOI: broker vs actual",
        ),
        VarianceFlagOut(
            field="broker_adr_growth_vs_market", rule_id="BROKER_VS_CBRE_ADR_GROWTH",
            severity="Warn", actual=0.03, broker=0.06, delta=0.03, delta_pct=0.03,
            note="Broker projects 6.0% Y1 ADR growth vs CBRE 3.0%",
        ),
        VarianceFlagOut(field="occupancy", severity="info"),
    ]
    rows = _variance_flags_from_out(flags)

    assert [r["flag_id"] for r in rows] == ["VF-001", "VF-002", "VF-003"]
    assert [r["severity"] for r in rows] == ["CRITICAL", "WARN", "INFO"]
    assert rows[0]["metric"] == "NOI"
    assert rows[0]["broker_value"] == 5_200_000.0
    assert rows[0]["t12_value"] == 4_181_000.0
    assert rows[0]["variance_pct"] == -0.196
    assert rows[0]["recommended_action"] == "NOI: broker vs actual"
    assert rows[0]["rule_id"] == "USALI-NOI-01"
    assert rows[1]["metric"] == "Broker ADR Growth vs Market"
    # Absent numbers are omitted — never written as 0.
    assert rows[2]["metric"] == "Occupancy"
    for k in ("broker_value", "t12_value", "variance_pct", "recommended_action"):
        assert k not in rows[2]


# ─────────────────── 6. demo/slug Kimpton path untouched ───────────────────


def test_demo_payload_unchanged() -> None:
    from app.export.fixtures import kimpton_deal, kimpton_memo, kimpton_model, load_demo_payload

    deal, model, memo = load_demo_payload("kimpton-angler-2026")
    assert model == kimpton_model()
    assert memo == kimpton_memo()
    expected_deal = kimpton_deal()
    expected_deal["id"] = "kimpton-angler-2026"
    assert deal == expected_deal
    assert len(model["scenario_outputs"]) == 3 and len(model["variance_flags"]) == 5
