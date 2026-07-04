"""Deterministic STR Trend template extraction (cost-opt pass W).

Covers the three recognized layouts plus the conservative-None
contract:

* legacy Custom Trend ``.xls`` (real golden-set fixture,
  ``tests/fixtures/sample_str_trend.xls``) — full positive path with
  exact pinned values (roster names/keys, 1124 total keys, subject
  name, monthly + annual series values read straight from the grid);
* modern Monthly STAR ``.xlsx`` — synthesized in-test with openpyxl to
  mirror the real ``ANG-20241200-USD-E.xlsx`` layout (same tab titles,
  same label anchors, same known-good numbers: 5 comps, 40+75+129+73+
  107 = 424 comp keys, subject TTM 83.2567 / 236.094 / 196.564,
  indices 112.059 / 79.1771 / 88.7253 on STR's 100-scale);
* modern Weekly STAR ``.xlsx`` — roster-only coverage, must NOT emit
  TTM metrics (the file has no trailing-twelve data);
* negatives — a real P&L xlsx, a non-STR doc_type, and the
  ``TEMPLATE_EXTRACTION_ENABLED=False`` passthrough in the
  ``documents.py`` wire-in helper.
"""

from __future__ import annotations

import io
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pytest

from app.extraction.parser import parse_document
from app.extraction.template_extractors import (
    TemplateExtractResult,
    try_template_extract,
)

FIXTURES = Path(__file__).parent / "fixtures"


async def _parse(path: Path):
    return await parse_document(
        file_bytes=path.read_bytes(), filename=path.name
    )


def _by_name(result: TemplateExtractResult) -> dict[str, Any]:
    return {f["field_name"]: f["value"] for f in result.fields}


# ── synthetic Monthly / Weekly STAR workbooks ────────────────────────
# Mirrors the layout of the real Kimpton Angler's exports (which live
# outside the repo and must not be committed): same tab titles, same
# label anchors, same values, so the assertions below pin the exact
# numbers the real ANG-20241200-USD-E.xlsx yields (424/556 keys).


def _monthly_star_xlsx_bytes() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    toc = wb.active
    toc.title = "Table of Contents"
    toc.cell(row=1, column=2, value="Table of Contents")

    header = [
        "Tab 2 - Monthly Performance at a Glance - My Property vs. "
        "Competitive Set - Performance Set",
        "Kimpton Angler's Hotel        600 Washington Ave        "
        "Miami Beach, FL 331396207",
        "Property ID: 56387        ChainID: 18033",
        "For the Month of: December 2024        Date Created: "
        "January 23, 2025",
    ]

    glance = wb.create_sheet("Glance_1")
    for r, text in enumerate(header, start=1):
        glance.cell(row=r, column=2, value=text)
    glance.cell(row=6, column=6, value="Occupancy (%)")
    glance.cell(row=6, column=11, value="ADR")
    glance.cell(row=6, column=16, value="RevPAR")
    for base in (7, 12, 17):
        glance.cell(row=7, column=base, value="My Prop")
        glance.cell(row=7, column=base + 1, value="Comp Set")
    glance.cell(row=7, column=9, value="Index (MPI)")
    glance.cell(row=7, column=14, value="Index (ARI)")
    glance.cell(row=7, column=19, value="Index (RGI)")
    rows = {
        "Current Month": (87.8299, 71.4853, 122.864, 278.767, 439.15,
                          63.4787, 244.841, 313.928, 77.9926),
        "Year To Date": (83.2567, 74.2971, 112.059, 236.094, 298.185,
                         79.1771, 196.564, 221.542, 88.7253),
        "Running 3 Month": (84.6756, 70.2847, 120.475, 224.387, 326.202,
                            68.7876, 190.001, 229.27, 82.872),
        "Running 12 Month": (83.2567, 74.2971, 112.059, 236.094, 298.185,
                             79.1771, 196.564, 221.542, 88.7253),
    }
    for r, (label, vals) in enumerate(rows.items(), start=8):
        glance.cell(row=r, column=4, value=label)
        for c, v in zip((7, 8, 9, 12, 13, 14, 17, 18, 19), vals):
            glance.cell(row=r, column=c, value=v)

    comp = wb.create_sheet("Comp_1")
    comp.cell(
        row=1, column=2,
        value="Tab 4 - Competitive Set Report - Performance Set",
    )
    for r, text in enumerate(header[1:], start=2):
        comp.cell(row=r, column=2, value=text)
    blocks = {
        "Occupancy (%)": (88.7879, 87.8299),
        "ADR": (197.741, 278.767),
        "RevPAR": (175.571, 244.841),
    }
    r = 5
    for label, (nov, dec) in blocks.items():
        comp.cell(row=r, column=2, value=label)
        comp.cell(row=r, column=3, value=2024)
        comp.cell(row=r + 1, column=3, value="Nov")
        comp.cell(row=r + 1, column=4, value="Dec")
        comp.cell(row=r + 2, column=2, value="My Property")
        comp.cell(row=r + 2, column=3, value=nov)
        comp.cell(row=r + 2, column=4, value=dec)
        r += 4

    resp = wb.create_sheet("Response_1")
    resp.cell(row=1, column=2, value="Tab 22 - Response Report - Performance Set")
    for r2, text in enumerate(header[1:], start=2):
        resp.cell(row=r2, column=2, value=text)
    roster_header = ["STR#", "Name", "City, State", "Zip", "Phone", "Rooms"]
    for c, label in enumerate(roster_header, start=3):
        resp.cell(row=6, column=c, value=label)
    roster = [
        (56387, "Kimpton Angler's Hotel", 132),
        (53909, "Z Ocean Hotel, Classico A Sonesta Collection", 40),
        (34401, "Blue Moon Hotel", 75),
        (44117, "The Betsy South Beach", 129),
        (39070, "The Tony Hotel of South Beach", 73),
        (33931, "Dream South Beach", 107),
    ]
    for r2, (sid, name, rooms) in enumerate(roster, start=7):
        resp.cell(row=r2, column=3, value=sid)
        resp.cell(row=r2, column=4, value=name)
        resp.cell(row=r2, column=5, value="Miami Beach, FL")
        resp.cell(row=r2, column=8, value=rooms)
    resp.cell(row=13, column=8, value=556)  # roster total incl. subject

    wb.create_sheet("Help").cell(row=1, column=1, value="Help")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _weekly_star_xlsx_bytes() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    toc = wb.active
    toc.title = "Table of Contents"
    toc.cell(row=1, column=2, value="Table of Contents")

    resp = wb.create_sheet("Response")
    resp.cell(row=1, column=2, value="Tab 14 - Response Report")
    resp.cell(row=2, column=3, value="Kimpton Angler's Hotel        600 Washington Ave")
    resp.cell(row=3, column=3, value="STR # 56387")
    resp.cell(row=4, column=3, value="For the Week of: May 25, 2025")
    for c, label in enumerate(
        ["STR ID", "Name", "City, State", "Zip", "Phone", "Rooms"], start=3
    ):
        resp.cell(row=6, column=c, value=label)
    roster = [
        (56387, "Kimpton Angler's Hotel", 132),
        (33931, "Dream South Beach", 107),
        (44117, "The Betsy South Beach", 129),
    ]
    for r, (sid, name, rooms) in enumerate(roster, start=7):
        resp.cell(row=r, column=3, value=sid)
        resp.cell(row=r, column=4, value=name)
        resp.cell(row=r, column=8, value=rooms)

    wb.create_sheet("Help").cell(row=1, column=1, value="Help")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── legacy .xls Custom Trend (real golden-set fixture) ───────────────


async def test_xls_custom_trend_positive() -> None:
    parsed = await _parse(FIXTURES / "sample_str_trend.xls")
    result = try_template_extract(parsed, "STR_TREND")
    assert result is not None
    assert result.template_name == "str_trend"
    assert "custom_trend_xls" in result.coverage_note

    values = _by_name(result)
    # Roster (Response tab) — names + keys pinned exactly.
    expected_comps = [
        ("The Ritz-Carlton South Beach", 376),
        ("The Setai", 150),
        ("Faena Hotel Miami Beach", 179),
        ("Ritz-Carlton Bal Harbour Miami", 102),
        ("St. Regis Bal Harbour Resort", 214),
        ("Four Seasons Hotel At The Surf Club", 103),
    ]
    for i, (name, keys) in enumerate(expected_comps, start=1):
        assert values[f"ttm_performance.compset.{i}.name"] == name
        assert values[f"ttm_performance.compset.{i}.keys"] == keys
    assert values["comp_set.comp_set_size"] == 6
    assert values["comp_set.total_keys"] == 1124  # matches "Total Properties" row

    assert values["ttm_performance.subject.name"] == (
        "Rosewood The Raleigh Miami Beach"
    )
    assert values["str_trend.report_year"] == 2023

    # Monthly series: most-recent 12 populated months x 3 metrics,
    # values read straight off the By Measure grid.
    monthly = [n for n in values if ".monthly." in n]
    assert len(monthly) == 36
    assert values[
        "ttm_performance.subject.monthly.2023_02.occupancy_pct"
    ] == pytest.approx(72.8584)
    assert values[
        "ttm_performance.subject.monthly.2023_02.adr_usd"
    ] == pytest.approx(1537.71)
    # Annual rollups from the Total Year column.
    assert values[
        "ttm_performance.subject.annual.2022.occupancy_pct"
    ] == pytest.approx(64.117)
    assert values[
        "ttm_performance.subject.annual.2022.adr_usd"
    ] == pytest.approx(1220.25)

    # Custom Trend publishes no subject-vs-comp-set indices or TTM row.
    assert "ttm_performance.indices.rgi_revpar_index" not in values
    assert "ttm_performance.subject.occupancy_pct" not in values


async def test_field_shape_matches_llm_extractor_contract() -> None:
    parsed = await _parse(FIXTURES / "sample_str_trend.xls")
    result = try_template_extract(parsed, "STR_TREND")
    assert result is not None
    for f in result.fields:
        assert set(f) >= {"field_name", "value", "confidence", "unit"}
        assert isinstance(f["field_name"], str) and f["field_name"]
        assert f["value"] is not None
        # Deterministically-read cells carry confidence 1.0.
        assert f["confidence"] == 1.0
        assert f["unit"] is None or isinstance(f["unit"], str)


# ── modern Monthly STAR .xlsx ────────────────────────────────────────


async def test_xlsx_monthly_star_positive() -> None:
    parsed = await parse_document(
        file_bytes=_monthly_star_xlsx_bytes(), filename="ANG-20241200-USD-E.xlsx"
    )
    result = try_template_extract(parsed, "STR_TREND")
    assert result is not None
    assert "monthly_star_xlsx" in result.coverage_note

    values = _by_name(result)
    # Comp-set roster: subject (Property ID 56387) excluded; the five
    # comps carry the known 40+75+129+73+107 = 424 keys.
    keys = [values[f"ttm_performance.compset.{i}.keys"] for i in range(1, 6)]
    assert keys == [40, 75, 129, 73, 107]
    assert values["comp_set.comp_set_size"] == 5
    assert values["comp_set.total_keys"] == 424
    assert "ttm_performance.compset.6.name" not in values  # subject skipped

    # Subject TTM = the Glance "Running 12 Month" row.
    assert values["ttm_performance.subject.occupancy_pct"] == pytest.approx(83.2567)
    assert values["ttm_performance.subject.adr_usd"] == pytest.approx(236.094)
    assert values["ttm_performance.subject.revpar_usd"] == pytest.approx(196.564)

    # Indices converted from STR's 100-scale to the canonical
    # 1.00 = parity scale.
    assert values["ttm_performance.indices.mpi_occupancy_index"] == pytest.approx(1.12059)
    assert values["ttm_performance.indices.ari_adr_index"] == pytest.approx(0.791771)
    assert values["ttm_performance.indices.rgi_revpar_index"] == pytest.approx(0.887253)

    assert values["ttm_performance.subject.name"] == "Kimpton Angler's Hotel"
    assert values["str_trend.report_year"] == 2024

    # Monthly series from the Comp tab.
    assert values[
        "ttm_performance.subject.monthly.2024_12.occupancy_pct"
    ] == pytest.approx(87.8299)
    assert values[
        "ttm_performance.subject.monthly.2024_11.adr_usd"
    ] == pytest.approx(197.741)
    assert values[
        "ttm_performance.subject.monthly.2024_12.revpar_usd"
    ] == pytest.approx(244.841)


async def test_xlsx_weekly_star_roster_only() -> None:
    parsed = await parse_document(
        file_bytes=_weekly_star_xlsx_bytes(), filename="56387-20250525-USD-E.xlsx"
    )
    result = try_template_extract(parsed, "STR_TREND")
    assert result is not None
    assert "weekly_star_xlsx" in result.coverage_note

    values = _by_name(result)
    assert values["comp_set.comp_set_size"] == 2
    assert values["comp_set.total_keys"] == 107 + 129
    assert values["str_trend.report_year"] == 2025
    # A weekly file has no trailing-twelve data: emitting TTM metrics
    # from it would be wrong, so they must be absent.
    assert "ttm_performance.subject.occupancy_pct" not in values
    assert "ttm_performance.indices.rgi_revpar_index" not in values
    assert not any(".monthly." in n for n in values)


# ── conservative-None contract ───────────────────────────────────────


async def test_pnl_xlsx_returns_none() -> None:
    parsed = await _parse(FIXTURES / "sam_anglers_2023_pnl.xlsx")
    assert try_template_extract(parsed, "STR_TREND") is None


async def test_non_str_doc_type_returns_none() -> None:
    parsed = await _parse(FIXTURES / "sample_str_trend.xls")
    assert try_template_extract(parsed, "T12") is None
    assert try_template_extract(parsed, "PNL_BENCHMARK") is None


async def test_str_doc_type_alias_accepted() -> None:
    # Legacy bare 'STR' label routes through the same template.
    parsed = await _parse(FIXTURES / "sample_str_trend.xls")
    assert try_template_extract(parsed, "STR") is not None


# ── wire-in helper: flag gating (documents.py) ───────────────────────


def _extraction_data_from(parsed) -> dict[str, Any]:
    return {
        "parser": parsed.parser,
        "total_pages": parsed.total_pages,
        "content_hash": parsed.content_hash,
        "pages": [
            {
                "page_num": p.page_num,
                "text": p.text,
                "tables": p.tables,
                "metadata": p.metadata,
            }
            for p in parsed.pages
        ],
    }


async def test_wirein_flag_off_passthrough(monkeypatch: pytest.MonkeyPatch) -> None:
    from app.api import documents as documents_module

    parsed = await _parse(FIXTURES / "sample_str_trend.xls")
    data = _extraction_data_from(parsed)

    monkeypatch.setattr(
        documents_module,
        "get_settings",
        lambda: SimpleNamespace(TEMPLATE_EXTRACTION_ENABLED=False),
    )
    assert (
        documents_module._try_template_extraction(
            extraction_data=data,
            filename="sample_str_trend.xls",
            doc_id="doc-1",
            doc_type="STR_TREND",
        )
        is None
    )

    # Flag on: same document, same helper → template hit.
    monkeypatch.setattr(
        documents_module,
        "get_settings",
        lambda: SimpleNamespace(TEMPLATE_EXTRACTION_ENABLED=True),
    )
    result = documents_module._try_template_extraction(
        extraction_data=data,
        filename="sample_str_trend.xls",
        doc_id="doc-1",
        doc_type="STR_TREND",
    )
    assert result is not None
    assert result.template_name == "str_trend"


async def test_wirein_non_str_doc_type_skips(monkeypatch: pytest.MonkeyPatch) -> None:
    from app.api import documents as documents_module

    parsed = await _parse(FIXTURES / "sample_str_trend.xls")
    monkeypatch.setattr(
        documents_module,
        "get_settings",
        lambda: SimpleNamespace(TEMPLATE_EXTRACTION_ENABLED=True),
    )
    assert (
        documents_module._try_template_extraction(
            extraction_data=_extraction_data_from(parsed),
            filename="sample_str_trend.xls",
            doc_id="doc-1",
            doc_type="T12",
        )
        is None
    )
